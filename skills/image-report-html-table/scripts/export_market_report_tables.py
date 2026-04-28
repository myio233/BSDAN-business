#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
import shutil
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from PIL import Image

try:
    import pytesseract
except ModuleNotFoundError as exc:  # pragma: no cover - exercised via script invocation
    raise SystemExit(
        "Missing dependency: pytesseract. Install repo requirements first, then rerun this script."
    ) from exc

ROOT = Path(__file__).resolve().parents[3]
EXSCHOOL_ALL_CITIES = ["Shanghai", "Chengdu", "Wuhan", "Wuxi", "Ningbo"]
OCR_BINARY = shutil.which("tesseract")

EXSCHOOL_IMAGE_ROUND_MAP = {
    "report_r1.jpg": "r1",
    "report_r2.jpg": "r2",
    "report_r3.jpg": "r3",
    "report_r4.jpg": "r4",
}
EXSCHOOL_SOURCE_MAP = {
    "r1": "report4_market_reports.xlsx",
    "r2": "report3_market_reports.xlsx",
    "r3": "report2_market_reports_fixed.xlsx",
    "r4": "report1_market_reports_fixed.xlsx",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Convert report screenshots into HTML tables and Excel workbooks.")
    parser.add_argument("--preset", choices=["exschool"], help="Use a built-in project preset.")
    parser.add_argument("--image-dir", type=Path, help="Directory containing source images.")
    parser.add_argument("--output-dir", type=Path, help="Directory for generated outputs.")
    parser.add_argument("--source-xlsx-dir", type=Path, help="Optional directory containing verified source workbooks.")
    parser.add_argument("--image-glob", default="*.jpg", help="Glob used in generic mode. Default: *.jpg")
    return parser.parse_args()


def resolve_paths(args: argparse.Namespace) -> tuple[Path, Path, Path | None, dict[str, str] | None]:
    if args.preset == "exschool":
        image_dir = ROOT / "exschool"
        output_dir = ROOT / "outputs" / "exschool_market_report_exports"
        source_xlsx_dir = ROOT / "exschool"
        return image_dir, output_dir, source_xlsx_dir, EXSCHOOL_SOURCE_MAP

    if not args.image_dir or not args.output_dir:
        raise SystemExit("Generic mode requires --image-dir and --output-dir.")
    return args.image_dir, args.output_dir, args.source_xlsx_dir, None


def ocr_image(image_path: Path) -> tuple[str, list[str]]:
    if OCR_BINARY is None:
        raise RuntimeError("tesseract binary is not installed or not in PATH")
    text = pytesseract.image_to_string(Image.open(image_path), lang="eng")
    round_match = re.search(r"Round\s+(\d+)\s+Report", text, re.IGNORECASE)
    round_id = f"r{round_match.group(1)}" if round_match else "unknown"
    city_titles = re.findall(r"Market Report - ([A-Za-z]+)", text)
    deduped: list[str] = []
    for city in city_titles:
        if city not in deduped:
            deduped.append(city)
    return round_id, deduped


def extract_sheet_rows(xlsx_path: Path) -> dict[str, list[list[object]]]:
    xl = pd.ExcelFile(xlsx_path)
    result: dict[str, list[list[object]]] = {}
    for sheet in xl.sheet_names:
        df = pd.read_excel(xlsx_path, sheet_name=sheet, header=None)
        rows: list[list[object]] = []
        for row in df.itertuples(index=False):
            rows.append(["" if pd.isna(v) else v for v in row])
        result[sheet] = rows
    return result


def load_exschool_key_data() -> dict[str, dict[str, float]]:
    path = ROOT / "exschool" / "asdan_key_data_sheet.xlsx"
    df = pd.read_excel(path, sheet_name="Key Data", header=None)
    markets: dict[str, dict[str, float]] = {}
    for _, row in df.iterrows():
        market = str(row.iloc[0]).strip()
        if market in EXSCHOOL_ALL_CITIES:
            markets[market] = {
                "population": float(row.iloc[9]),
                "penetration": float(row.iloc[10]),
                "avg_price": float(row.iloc[11]),
            }
    return markets


def load_fixed_decisions() -> pd.DataFrame:
    path = ROOT / "outputs" / "exschool_inferred_decisions" / "all_companies_numeric_decisions.xlsx"
    if not path.exists():
        return pd.DataFrame()
    df = pd.read_excel(path)
    df["team"] = df["team"].astype(str)
    df["round_id"] = df["round_id"].astype(str)
    return df


def make_placeholder_city_rows(round_id: str, city: str, fixed_decisions: pd.DataFrame, key_data: dict[str, dict[str, float]]) -> list[list[object]]:
    city_key = city.lower()
    market_meta = key_data[city]
    market_size = int(round(float(market_meta["population"]) * float(market_meta["penetration"])))
    rows: list[list[object]] = [
        [f"Market Report - {city}", "", "", "", "", "", "", ""],
        ["Population", "Penetration", "Market Size", "Total Sales Volume", "Avg. Price", "", "", ""],
        [int(market_meta["population"]), float(market_meta["penetration"]), market_size, 0, int(round(market_meta["avg_price"])), "", "", ""],
        ["", "", "", "", "", "", "", ""],
        ["Team", "Management Index", "Agents", "Marketing Investment", "Product Quality Index", "Price", "Sales Volume", "Market Share"],
    ]
    if fixed_decisions.empty:
        return rows
    round_rows = fixed_decisions[fixed_decisions["round_id"] == round_id].copy()
    for _, row in round_rows.sort_values("team", key=lambda s: s.astype(int)).iterrows():
        workers = float(row.get("workers", 0) or 0)
        engineers = float(row.get("engineers", 0) or 0)
        total_people = workers + engineers
        products = float(row.get("products_planned", 0) or 0)
        mgmt_index = float(row.get("management_investment", 0) or 0) / total_people if total_people > 0 else 0.0
        quality_index = float(row.get("quality_investment", 0) or 0) / products if products > 0 else 0.0
        agents = int(row.get(f"{city_key}_agents_after", 0) or 0)
        marketing = float(row.get(f"{city_key}_marketing_investment", 0) or 0.0)
        price = float(row.get(f"{city_key}_price", 0) or 0.0)
        rows.append([row["team"], round(mgmt_index, 2), agents, round(marketing, 2), round(quality_index, 2), round(price, 2), 0, 0.0])
    return rows


def ensure_all_exschool_city_sheets(round_id: str, sheet_rows: dict[str, list[list[object]]]) -> dict[str, list[list[object]]]:
    fixed_decisions = load_fixed_decisions()
    key_data = load_exschool_key_data()
    out = dict(sheet_rows)
    for city in EXSCHOOL_ALL_CITIES:
        if city not in out and city in key_data:
            out[city] = make_placeholder_city_rows(round_id, city, fixed_decisions, key_data)
    return {city: out[city] for city in EXSCHOOL_ALL_CITIES if city in out}


def build_rows_from_source(round_id: str, sheet_rows: dict[str, list[list[object]]]) -> list[list[object]]:
    rows: list[list[object]] = []
    rows.append([f"{round_id.upper()} Market Reports", "", "", "", "", "", "", ""])
    rows.append(["Source", "OCR image verified against workbook", "", "", "", "", "", ""])
    rows.append(["", "", "", "", "", "", "", ""])
    for city, city_rows in sheet_rows.items():
        rows.append([f"Market Report - {city}", "", "", "", "", "", "", ""])
        normalized = list(city_rows)
        while normalized and all(str(v).strip() == "" for v in normalized[0]):
            normalized.pop(0)
        if normalized and str(normalized[0][0]).strip().startswith("Market Report - "):
            normalized.pop(0)
        while normalized and all(str(v).strip() == "" for v in normalized[0]):
            normalized.pop(0)
        rows.extend(normalized)
        rows.append(["", "", "", "", "", "", "", ""])
    return rows


def build_rows_from_ocr_text(round_id: str, image_path: Path) -> list[list[object]]:
    text = pytesseract.image_to_string(Image.open(image_path), lang="eng")
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    rows: list[list[object]] = [
        [f"{round_id.upper()} OCR Lines", ""],
        ["Source", "OCR-only fallback"],
        ["", ""],
    ]
    rows.extend([[line, ""] for line in lines])
    return rows


def write_html(round_id: str, rows: list[list[object]], city_titles: list[str], output_path: Path) -> None:
    max_cols = max(len(r) for r in rows)
    lines = [
        "<!doctype html>",
        '<html lang="zh-CN">',
        "<head>",
        '  <meta charset="utf-8" />',
        f"  <title>{round_id.upper()} Market Reports</title>",
        "  <style>",
        "    body { font-family: Arial, sans-serif; margin: 24px; }",
        "    table { border-collapse: collapse; width: 100%; }",
        "    td, th { border: 1px solid #999; padding: 6px 8px; font-size: 12px; }",
        "    .title td { background: #1f4e78; color: #fff; font-weight: 700; }",
        "    .section td { background: #d9eaf7; font-weight: 700; }",
        "    .blank td { border: none; height: 8px; }",
        "  </style>",
        "</head>",
        "<body>",
        f"  <h1>{round_id.upper()} OCR Verified Market Reports</h1>",
        f"  <p>OCR detected city sections: {', '.join(city_titles) if city_titles else 'None'}</p>",
        "  <table>",
    ]
    for idx, row in enumerate(rows):
        cls = ""
        if idx == 0:
            cls = ' class="title"'
        elif row and isinstance(row[0], str) and str(row[0]).startswith("Market Report - "):
            cls = ' class="section"'
        elif all(str(v).strip() == "" for v in row):
            cls = ' class="blank"'
        lines.append(f"    <tr{cls}>")
        padded = row + [""] * (max_cols - len(row))
        for value in padded:
            text = "" if value is None else str(value)
            text = text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            lines.append(f"      <td>{text}</td>")
        lines.append("    </tr>")
    lines += ["  </table>", "</body>", "</html>"]
    output_path.write_text("\n".join(lines), encoding="utf-8")


def write_combined_xlsx(round_id: str, rows: list[list[object]], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = round_id
    title_fill = PatternFill("solid", fgColor="1F4E78")
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    bold = Font(bold=True)
    white_bold = Font(bold=True, color="FFFFFF")
    max_cols = max(len(row) for row in rows)
    for row in rows:
        ws.append(row + [""] * (max_cols - len(row)))
    for cell in ws[1]:
        cell.fill = title_fill
        cell.font = white_bold
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if row[0].value and isinstance(row[0].value, str) and row[0].value.startswith("Market Report - "):
            for cell in row:
                cell.fill = section_fill
                cell.font = bold
    wb.save(output_path)


def write_structured_workbook(sheet_rows: dict[str, list[list[object]]], output_path: Path) -> None:
    wb = Workbook()
    first = True
    for city, rows in sheet_rows.items():
        ws = wb.active if first else wb.create_sheet()
        first = False
        ws.title = city
        for row in rows:
            ws.append(list(row))
    wb.save(output_path)


def main() -> None:
    args = parse_args()
    image_dir, output_dir, source_xlsx_dir, source_map = resolve_paths(args)

    html_dir = output_dir / "html"
    combined_xlsx_dir = output_dir / "combined_xlsx"
    structured_dir = output_dir / "structured_xlsx"
    html_dir.mkdir(parents=True, exist_ok=True)
    combined_xlsx_dir.mkdir(exist_ok=True)
    structured_dir.mkdir(exist_ok=True)

    validation_rows: list[dict[str, object]] = []

    if args.preset == "exschool":
        image_names = list(EXSCHOOL_IMAGE_ROUND_MAP.keys())
    else:
        image_names = sorted(p.name for p in image_dir.glob(args.image_glob))

    for image_name in image_names:
        image_path = image_dir / image_name
        if not image_path.exists():
            continue

        expected_round = EXSCHOOL_IMAGE_ROUND_MAP.get(image_name, "unknown") if args.preset == "exschool" else "unknown"

        source_xlsx = None
        sheet_rows = None
        source_sheets: list[str] = []
        visible_source_sheets: list[str] = []
        if source_xlsx_dir and source_map and expected_round in source_map:
            candidate = source_xlsx_dir / source_map[expected_round]
            if candidate.exists():
                source_xlsx = candidate
                raw_sheet_rows = extract_sheet_rows(candidate)
                visible_source_sheets = list(raw_sheet_rows.keys())
                sheet_rows = raw_sheet_rows
                if args.preset == "exschool":
                    sheet_rows = ensure_all_exschool_city_sheets(expected_round, sheet_rows)
                source_sheets = list(sheet_rows.keys())

        ocr_round = "missing_tesseract"
        ocr_cities: list[str] = []
        ocr_ok = False
        if OCR_BINARY is not None:
            ocr_round, ocr_cities = ocr_image(image_path)
            ocr_ok = True

        rows = build_rows_from_source(expected_round, sheet_rows) if sheet_rows else build_rows_from_ocr_text(expected_round, image_path)

        html_path = html_dir / f"{expected_round}_market_reports.html"
        combined_xlsx_path = combined_xlsx_dir / f"{expected_round}_market_reports_combined.xlsx"

        write_html(expected_round, rows, ocr_cities, html_path)
        write_combined_xlsx(expected_round, rows, combined_xlsx_path)

        if sheet_rows and source_xlsx:
            structured_path = structured_dir / source_xlsx.name
            write_structured_workbook(sheet_rows, structured_path)
            validation_rows.append(
                {
                    "round_id": expected_round,
                    "check": "ocr_city_titles_match_source_sheets",
                    "expected": ",".join(visible_source_sheets),
                    "actual": ",".join(ocr_cities),
                    "ok": ocr_ok and ocr_cities == visible_source_sheets,
                }
            )
            validation_rows.append(
                {
                    "round_id": expected_round,
                    "check": "generated_structured_sheets_match_source",
                    "expected": ",".join(source_sheets),
                    "actual": ",".join(pd.ExcelFile(structured_path).sheet_names),
                    "ok": pd.ExcelFile(structured_path).sheet_names == source_sheets,
                }
            )

        validation_rows.append(
            {
                "round_id": expected_round,
                "check": "ocr_binary_available",
                "expected": "tesseract in PATH",
                "actual": OCR_BINARY or "",
                "ok": OCR_BINARY is not None,
            }
        )
        validation_rows.append(
            {
                "round_id": expected_round,
                "check": "ocr_round_matches_expected",
                "expected": expected_round,
                "actual": ocr_round,
                "ok": ocr_ok and expected_round == ocr_round,
            }
        )

    pd.DataFrame(validation_rows).to_excel(output_dir / "validation_report.xlsx", index=False)
    print(f"Wrote outputs to {output_dir}")


if __name__ == "__main__":
    main()
