#!/usr/bin/env python3
import os
import re
import time
from PIL import Image as PILImage
from azure.ai.formrecognizer import DocumentAnalysisClient
from azure.core.credentials import AzureKeyCredential
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def get_azure_client():
    endpoint = os.environ.get("DOCUMENTINTELLIGENCE_ENDPOINT")
    key = os.environ.get("DOCUMENTINTELLIGENCE_API_KEY")
    return DocumentAnalysisClient(
        endpoint=endpoint,
        credential=AzureKeyCredential(key)
    )


def split_image_vertically(image_path, max_height=3000, overlap=200):
    """Split tall image into overlapping vertical chunks"""
    img = PILImage.open(image_path)
    print(f"  Original size: {img.size}")

    chunks = []
    y = 0
    chunk_num = 0

    while y < img.height:
        chunk_height = min(max_height, img.height - y)
        box = (0, y, img.width, y + chunk_height)
        chunk = img.crop(box)

        chunk_path = image_path + f".chunk{chunk_num}.jpg"
        chunk.save(chunk_path, "JPEG", quality=80)
        chunks.append(chunk_path)

        y += chunk_height - overlap
        chunk_num += 1

    print(f"  Split into {len(chunks)} chunks")
    return chunks


def merge_related_tables(tables):
    if len(tables) <= 1:
        return tables

    merged = []
    i = 0
    while i < len(tables):
        table = tables[i]
        if i + 1 < len(tables) and len(table) <= 3:
            next_table = tables[i + 1]
            if len(table[0]) == len(next_table[0]) or abs(len(table[0]) - len(next_table[0])) <= 2:
                merged_table = table + next_table
                merged.append(merged_table)
                i += 2
                continue
        merged.append(table)
        i += 1
    return merged


def analyze_chunk(client, chunk_path):
    """Analyze a single chunk"""
    with open(chunk_path, "rb") as f:
        poller = client.begin_analyze_document("prebuilt-layout", document=f)
    result = poller.result()

    tables = []
    for table in result.tables:
        rows = [["" for _ in range(table.column_count)] for _ in range(table.row_count)]
        for cell in table.cells:
            rows[cell.row_index][cell.column_index] = cell.content
        tables.append(rows)

    full_text = ""
    for page in result.pages:
        for line in page.lines:
            full_text += line.content + "\n"

    return tables, full_text


def save_table_to_excel(table_data, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Market Report"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for row_idx, row_data in enumerate(table_data):
        ws.append(row_data)
        row_str = " ".join(str(cell) for cell in row_data)
        is_header = ("Population" in row_str and "Penetration" in row_str) or \
                    ("Team" in row_str and ("Management" in row_str or "Index" in row_str))
        if is_header:
            for cell in ws[ws.max_row]:
                cell.fill = header_fill
                cell.font = header_font

    for column in ws.columns:
        max_len = 0
        col_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 55)

    for attempt in range(3):
        try:
            if os.path.exists(output_path):
                os.remove(output_path)
                time.sleep(0.1)
            wb.save(output_path)
            return True
        except:
            if attempt == 2:
                base, ext = os.path.splitext(output_path)
                output_path = f"{base}_v3{ext}"
                wb.save(output_path)
                return True
            time.sleep(0.2)


def process_r7():
    folder_path = "/mnt/c/Users/david/documents/ASDAN/表格/WYEF/r7"
    folder_name = "r7"
    print(f"Processing: {folder_name}")

    market_img = os.path.join(folder_path, "market_report.png")

    print("  Splitting image...")
    chunk_paths = split_image_vertically(market_img)

    client = get_azure_client()

    all_tables = []
    all_text = ""

    for i, chunk_path in enumerate(chunk_paths):
        print(f"  Analyzing chunk {i+1}/{len(chunk_paths)}...")
        tables, text = analyze_chunk(client, chunk_path)
        all_tables.extend(tables)
        all_text += text

        os.remove(chunk_path)

    all_tables = merge_related_tables(all_tables)
    print(f"  Found {len(all_tables)} tables total")

    # Find cities
    lines = all_text.split('\n')
    cities = []
    current_city = None
    start_idx = 0
    for i, line in enumerate(lines):
        if "Market Report - " in line:
            if current_city:
                cities.append((current_city, start_idx, i))
            current_city = line.replace("Market Report - ", "").strip()
            start_idx = i
    if current_city:
        cities.append((current_city, start_idx, len(lines)))
    if not cities:
        cities = [("Market", 0, len(lines))]

    print(f"  Found {len(cities)} cities")

    # Save tables
    for i, table in enumerate(all_tables):
        if i < len(cities):
            city_name = cities[i][0]
        else:
            city_name = f"Table_{i+1}"

        if len(table) < 2:
            continue

        safe_name = re.sub(r'[^\w\s-]', '', city_name).replace(' ', '_')
        filename = f"{folder_name}_market_{safe_name}_azure.xlsx"
        filepath = os.path.join(folder_path, filename)

        if save_table_to_excel(table, filepath):
            print(f"    Saved: {filename} ({len(table)} rows)")


if __name__ == "__main__":
    process_r7()
