#!/usr/bin/env python3
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os

# Define all the data from the images
data = {
    # r4 folder - earliest round
    "r4_market_shanghai": {
        "market_info": {
            "Population": 6000000,
            "Penetration": "2.66%",
            "Market Size": 159600,
            "Total Sales Volume": 152388,
            "Avg. Price": 17908
        },
        "teams": [
            {"Team": 2, "Management Index": 0, "Agents": 1, "Marketing Investment": 0, "Product Quality Index": 0, "Price": 24688, "Sales Volume": 0, "Market Share": "0.00%"},
            {"Team": 3, "Management Index": 0, "Agents": 4, "Marketing Investment": 0, "Product Quality Index": 0, "Price": 17410, "Sales Volume": 1069, "Market Share": "0.66%"},
            {"Team": 4, "Management Index": 0, "Agents": 9, "Marketing Investment": 20000000, "Product Quality Index": 729.33, "Price": 20999, "Sales Volume": 19017, "Market Share": "11.91%"},
            {"Team": 5, "Management Index": 0, "Agents": 1, "Marketing Investment": 0, "Product Quality Index": 0, "Price": 21000, "Sales Volume": 83, "Market Share": "0.05%"},
            {"Team": 6, "Management Index": 309.38, "Agents": 6, "Marketing Investment": 28000000, "Product Quality Index": 250.00, "Price": 22677, "Sales Volume": 12507, "Market Share": "7.83%"},
            {"Team": 7, "Management Index": 0, "Agents": 3, "Marketing Investment": 0, "Product Quality Index": 0, "Price": 17222, "Sales Volume": 4241, "Market Share": "2.65%"},
            {"Team": 8, "Management Index": 0, "Agents": 2, "Marketing Investment": 0, "Product Quality Index": 0.02, "Price": 23999, "Sales Volume": 80, "Market Share": "0.05%"},
            {"Team": 9, "Management Index": 3104.20, "Agents": 10, "Marketing Investment": 50000000, "Product Quality Index": 10.00, "Price": 21111, "Sales Volume": 25134, "Market Share": "15.74%"},
            {"Team": 10, "Management Index": 0, "Agents": 1, "Marketing Investment": 44444, "Product Quality Index": 0, "Price": 15777, "Sales Volume": 68693, "Market Share": "43.04%"},
            {"Team": 11, "Management Index": 3.52, "Agents": 1, "Marketing Investment": 520, "Product Quality Index": 5.32, "Price": 14999, "Sales Volume": 179, "Market Share": "0.11%"},
            {"Team": 12, "Management Index": 1303.31, "Agents": 7, "Marketing Investment": 8000000, "Product Quality Index": 814.64, "Price": 21088, "Sales Volume": 21385, "Market Share": "13.39%"},
        ]
    },
    "r4_market_chengdu": {
        "market_info": {
            "Population": 4000000,
            "Penetration": "2.12%",
            "Market Size": 84800,
            "Total Sales Volume": 81910,
            "Avg. Price": 18083
        },
        "teams": [
            {"Team": 3, "Management Index": 0, "Agents": 3, "Marketing Investment": 0, "Product Quality Index": 0, "Price": 15980, "Sales Volume": 9443, "Market Share": "11.13%"},
            {"Team": 4, "Management Index": 0, "Agents": 5, "Marketing Investment": 12000000, "Product Quality Index": 729.33, "Price": 20999, "Sales Volume": 11894, "Market Share": "14.02%"},
            {"Team": 5, "Management Index": 0, "Agents": 4, "Marketing Investment": 0, "Product Quality Index": 0, "Price": 20500, "Sales Volume": 11, "Market Share": "0.01%"},
            {"Team": 6, "Management Index": 309.38, "Agents": 3, "Marketing Investment": 0, "Product Quality Index": 250.00, "Price": 22322, "Sales Volume": 2650, "Market Share": "3.12%"},
            {"Team": 7, "Management Index": 0, "Agents": 1, "Marketing Investment": 0, "Product Quality Index": 0, "Price": 15888, "Sales Volume": 27419, "Market Share": "32.33%"},
            {"Team": 8, "Management Index": 0, "Agents": 2, "Marketing Investment": 0, "Product Quality Index": 0.02, "Price": 23999, "Sales Volume": 21, "Market Share": "0.02%"},
            {"Team": 9, "Management Index": 3104.20, "Agents": 8, "Marketing Investment": 50000000, "Product Quality Index": 10.00, "Price": 21111, "Sales Volume": 16312, "Market Share": "19.23%"},
            {"Team": 10, "Management Index": 0, "Agents": 10, "Marketing Investment": 19777777, "Product Quality Index": 0, "Price": 21977, "Sales Volume": 6976, "Market Share": "8.22%"},
            {"Team": 12, "Management Index": 1303.31, "Agents": 4, "Marketing Investment": 4120000, "Product Quality Index": 814.64, "Price": 21088, "Sales Volume": 7184, "Market Share": "8.47%"},
        ]
    },
    "r4_market_hangzhou": {
        "market_info": {
            "Population": 2500000,
            "Penetration": "1.72%",
            "Market Size": 43000,
            "Total Sales Volume": 26209,
            "Avg. Price": 19811
        },
        "teams": [
            {"Team": 2, "Management Index": 0, "Agents": 1, "Marketing Investment": 0, "Product Quality Index": 0, "Price": 24688, "Sales Volume": 0, "Market Share": "0.00%"},
            {"Team": 3, "Management Index": 0, "Agents": 1, "Marketing Investment": 0, "Product Quality Index": 0, "Price": 20000, "Sales Volume": 223, "Market Share": "0.51%"},
            {"Team": 4, "Management Index": 0, "Agents": 3, "Marketing Investment": 8000000, "Product Quality Index": 729.33, "Price": 19999, "Sales Volume": 8096, "Market Share": "18.82%"},
            {"Team": 5, "Management Index": 0, "Agents": 1, "Marketing Investment": 0, "Product Quality Index": 0, "Price": 21000, "Sales Volume": 43, "Market Share": "0.10%"},
            {"Team": 6, "Management Index": 309.38, "Agents": 3, "Marketing Investment": 0, "Product Quality Index": 250.00, "Price": 22322, "Sales Volume": 2650, "Market Share": "0.52%"},
            {"Team": 8, "Management Index": 0, "Agents": 2, "Marketing Investment": 0, "Product Quality Index": 0.02, "Price": 24999, "Sales Volume": 8, "Market Share": "0.01%"},
            {"Team": 9, "Management Index": 3104.20, "Agents": 8, "Marketing Investment": 20000000, "Product Quality Index": 10.00, "Price": 22222, "Sales Volume": 11554, "Market Share": "26.86%"},
            {"Team": 10, "Management Index": 0, "Agents": 10, "Marketing Investment": 7771288, "Product Quality Index": 0, "Price": 21977, "Sales Volume": 2236, "Market Share": "5.20%"},
            {"Team": 11, "Management Index": 3.52, "Agents": 1, "Marketing Investment": 520, "Product Quality Index": 5.32, "Price": 14999, "Sales Volume": 64, "Market Share": "0.14%"},
            {"Team": 12, "Management Index": 1303.31, "Agents": 7, "Marketing Investment": 880000, "Product Quality Index": 814.64, "Price": 20399, "Sales Volume": 3758, "Market Share": "8.73%"},
        ]
    },
    "r4_sales": {
        "agents": [
            {"Market": "Shanghai", "Previous": 7, "Change": "+3", "After": 10, "Change Cost": 900000, "Marketing Investment": 50000000},
            {"Market": "Chengdu", "Previous": 5, "Change": "+3", "After": 8, "Change Cost": 900000, "Marketing Investment": 50000000},
            {"Market": "Hangzhou", "Previous": 5, "Change": "+3", "After": 8, "Change Cost": 900000, "Marketing Investment": 20000000},
        ],
        "market_sales": [
            {"Market": "Shanghai", "Competitive Power": "27.79%", "Sales Volume": 25134, "Market Share": "15.74%", "Price": 21111, "Sales": 530603874},
            {"Market": "Chengdu", "Competitive Power": "29.07%", "Sales Volume": 16312, "Market Share": "19.23%", "Price": 21111, "Sales": 344362632},
            {"Market": "Hangzhou", "Competitive Power": "26.86%", "Sales Volume": 11554, "Market Share": "26.86%", "Price": 22222, "Sales": 256752988},
        ]
    },

    # r3 folder
    "r3_market_shanghai": {
        "market_info": {
            "Population": 6000000,
            "Penetration": "2.42%",
            "Market Size": 145200,
            "Total Sales Volume": 78244,
            "Avg. Price": 17418
        },
        "teams": [
            {"Team": 2, "Management Index": 0, "Agents": 1, "Marketing Investment": 0, "Product Quality Index": 0, "Price": 24688, "Sales Volume": 0, "Market Share": "0.00%"},
            {"Team": 3, "Management Index": 0, "Agents": 4, "Marketing Investment": 0, "Product Quality Index": 0, "Price": 21000, "Sales Volume": 93, "Market Share": "0.06%"},
            {"Team": 4, "Management Index": 0, "Agents": 6, "Marketing Investment": 11000000, "Product Quality Index": 0, "Price": 23777, "Sales Volume": 14674, "Market Share": "10.10%"},
            {"Team": 5, "Management Index": 1999.88, "Agents": 1, "Marketing Investment": 388, "Product Quality Index": 0.98, "Price": 24366, "Sales Volume": 3208, "Market Share": "2.20%"},
            {"Team": 6, "Management Index": 11.14, "Agents": 4, "Marketing Investment": 12000000, "Product Quality Index": 1400.00, "Price": 23444, "Sales Volume": 14300, "Market Share": "9.84%"},
            {"Team": 7, "Management Index": 2867.77, "Agents": 3, "Marketing Investment": 0, "Product Quality Index": 0, "Price": 23777, "Sales Volume": 12537, "Market Share": "8.63%"},
            {"Team": 8, "Management Index": 1393.52, "Agents": 2, "Marketing Investment": 0, "Product Quality Index": 0.95, "Price": 24999, "Sales Volume": 199, "Market Share": "0.13%"},
            {"Team": 9, "Management Index": 10.00, "Agents": 7, "Marketing Investment": 16000000, "Product Quality Index": 10.00, "Price": 22222, "Sales Volume": 10292, "Market Share": "7.08%"},
            {"Team": 10, "Management Index": 0, "Agents": 1, "Marketing Investment": 7777, "Product Quality Index": 981.49, "Price": 24277, "Sales Volume": 8786, "Market Share": "6.05%"},
            {"Team": 11, "Management Index": 0, "Agents": 1, "Marketing Investment": 0, "Product Quality Index": 0, "Price": 12000, "Sales Volume": 2204, "Market Share": "1.51%"},
            {"Team": 12, "Management Index": 1303.31, "Agents": 4, "Marketing Investment": 0, "Product Quality Index": 596.25, "Price": 24088, "Sales Volume": 11951, "Market Share": "8.23%"},
        ]
    },
    "r3_market_chengdu": {
        "market_info": {
            "Population": 4000000,
            "Penetration": "1.93%",
            "Market Size": 77200,
            "Total Sales Volume": 27850,
            "Avg. Price": 15993
        },
        "teams": [
            {"Team": 3, "Management Index": 0, "Agents": 3, "Marketing Investment": 0, "Product Quality Index": 0, "Price": 21000, "Sales Volume": 9, "Market Share": "0.01%"},
            {"Team": 4, "Management Index": 0, "Agents": 2, "Marketing Investment": 0, "Product Quality Index": 0, "Price": 23777, "Sales Volume": 0, "Market Share": "0.00%"},
            {"Team": 5, "Management Index": 1999.88, "Agents": 4, "Marketing Investment": 8000888, "Product Quality Index": 0.98, "Price": 24346, "Sales Volume": 2470, "Market Share": "3.19%"},
            {"Team": 6, "Management Index": 11.14, "Agents": 1, "Marketing Investment": 0, "Product Quality Index": 1400.00, "Price": 23222, "Sales Volume": 3575, "Market Share": "4.63%"},
            {"Team": 7, "Management Index": 2867.77, "Agents": 1, "Marketing Investment": 0, "Product Quality Index": 0, "Price": 23777, "Sales Volume": 4178, "Market Share": "5.41%"},
            {"Team": 8, "Management Index": 1393.52, "Agents": 2, "Marketing Investment": 0, "Product Quality Index": 0.95, "Price": 24999, "Sales Volume": 79, "Market Share": "0.10%"},
            {"Team": 9, "Management Index": 10.00, "Agents": 5, "Marketing Investment": 13000000, "Product Quality Index": 10.00, "Price": 22222, "Sales Volume": 7060, "Market Share": "9.14%"},
            {"Team": 10, "Management Index": 0, "Agents": 7, "Marketing Investment": 5047777, "Product Quality Index": 981.49, "Price": 24277, "Sales Volume": 9423, "Market Share": "12.20%"},
            {"Team": 12, "Management Index": 1303.31, "Agents": 1, "Marketing Investment": 0, "Product Quality Index": 596.25, "Price": 24088, "Sales Volume": 1056, "Market Share": "1.36%"},
        ]
    },
    "r3_market_hangzhou": {
        "market_info": {
            "Population": 2500000,
            "Penetration": "1.57%",
            "Market Size": 39250,
            "Total Sales Volume": 20960,
            "Avg. Price": 18351
        },
        "teams": [
            {"Team": 2, "Management Index": 0, "Agents": 1, "Marketing Investment": 0, "Product Quality Index": 0, "Price": 24688, "Sales Volume": 0, "Market Share": "0.00%"},
            {"Team": 3, "Management Index": 0, "Agents": 1, "Marketing Investment": 0, "Product Quality Index": 0, "Price": 21000, "Sales Volume": 8, "Market Share": "0.02%"},
            {"Team": 5, "Management Index": 1999.88, "Agents": 1, "Marketing Investment": 888, "Product Quality Index": 0.98, "Price": 24366, "Sales Volume": 5292, "Market Share": "13.48%"},
            {"Team": 6, "Management Index": 11.14, "Agents": 1, "Marketing Investment": 8888, "Product Quality Index": 730.00, "Price": 23222, "Sales Volume": 2802, "Market Share": "9.10%"},
            {"Team": 8, "Management Index": 2503.12, "Agents": 2, "Marketing Investment": 0, "Product Quality Index": 0, "Price": 22999, "Sales Volume": 3866, "Market Share": "10.81%"},
            {"Team": 9, "Management Index": 10.00, "Agents": 5, "Marketing Investment": 3000000, "Product Quality Index": 10.00, "Price": 21999, "Sales Volume": 6648, "Market Share": "16.93%"},
            {"Team": 10, "Management Index": 0, "Agents": 4, "Marketing Investment": 3047777, "Product Quality Index": 981.49, "Price": 24277, "Sales Volume": 3748, "Market Share": "10.48%"},
            {"Team": 12, "Management Index": 0, "Agents": 1, "Marketing Investment": 2000000, "Product Quality Index": 596.25, "Price": 24088, "Sales Volume": 355, "Market Share": "0.90%"},
        ]
    },
    "r3_sales": {
        "agents": [
            {"Market": "Shanghai", "Previous": 4, "Change": "+3", "After": 7, "Change Cost": 900000, "Marketing Investment": 16000000},
            {"Market": "Chengdu", "Previous": 2, "Change": "+3", "After": 5, "Change Cost": 900000, "Marketing Investment": 13000000},
            {"Market": "Hangzhou", "Previous": 2, "Change": "+3", "After": 5, "Change Cost": 900000, "Marketing Investment": 3000000},
        ],
        "market_sales": [
            {"Market": "Shanghai", "Competitive Power": "16.75%", "Sales Volume": 10292, "Market Share": "7.08%", "Price": 22222, "Sales": 228708824},
            {"Market": "Chengdu", "Competitive Power": "18.33%", "Sales Volume": 7060, "Market Share": "9.14%", "Price": 22222, "Sales": 156887320},
            {"Market": "Hangzhou", "Competitive Power": "16.93%", "Sales Volume": 6648, "Market Share": "16.93%", "Price": 21999, "Sales": 146249352},
        ]
    },

    # r2 folder
    "r2_market_shanghai": {
        "market_info": {
            "Population": 6000000,
            "Penetration": "2.20%",
            "Market Size": 132000,
            "Total Sales Volume": 23321,
            "Avg. Price": 13008
        },
        "teams": [
            {"Team": 2, "Management Index": 0, "Agents": 1, "Marketing Investment": 0, "Product Quality Index": 0, "Price": 24688, "Sales Volume": 0, "Market Share": "0.00%"},
            {"Team": 3, "Management Index": 0, "Agents": 4, "Marketing Investment": 500, "Product Quality Index": 0.62, "Price": 24444, "Sales Volume": 93, "Market Share": "0.07%"},
            {"Team": 4, "Management Index": 0, "Agents": 3, "Marketing Investment": 4748142, "Product Quality Index": 0, "Price": 24444, "Sales Volume": 2693, "Market Share": "2.04%"},
            {"Team": 5, "Management Index": 2346.02, "Agents": 1, "Marketing Investment": 388, "Product Quality Index": 6.97, "Price": 24388, "Sales Volume": 1324, "Market Share": "1.00%"},
            {"Team": 6, "Management Index": 11.45, "Agents": 1, "Marketing Investment": 8888, "Product Quality Index": 730.00, "Price": 24200, "Sales Volume": 2802, "Market Share": "2.12%"},
            {"Team": 7, "Management Index": 0, "Agents": 3, "Marketing Investment": 5000000, "Product Quality Index": 0, "Price": 24444, "Sales Volume": 3809, "Market Share": "2.88%"},
            {"Team": 8, "Management Index": 2503.12, "Agents": 1, "Marketing Investment": 0, "Product Quality Index": 0, "Price": 24999, "Sales Volume": 92, "Market Share": "0.06%"},
            {"Team": 9, "Management Index": 2.01, "Agents": 4, "Marketing Investment": 6000000, "Product Quality Index": 10.00, "Price": 24333, "Sales Volume": 7778, "Market Share": "5.89%"},
            {"Team": 11, "Management Index": 0, "Agents": 1, "Marketing Investment": 0, "Product Quality Index": 0, "Price": 23888, "Sales Volume": 0, "Market Share": "0.00%"},
            {"Team": 12, "Management Index": 0, "Agents": 1, "Marketing Investment": 0, "Product Quality Index": 568.92, "Price": 24333, "Sales Volume": 4730, "Market Share": "3.58%"},
        ]
    },
    "r2_market_chengdu": {
        "market_info": {
            "Population": 4000000,
            "Penetration": "1.76%",
            "Market Size": 70400,
            "Total Sales Volume": 10495,
            "Avg. Price": 12527
        },
        "teams": [
            {"Team": 3, "Management Index": 0, "Agents": 3, "Marketing Investment": 2500000, "Product Quality Index": 0.62, "Price": 24444, "Sales Volume": 1379, "Market Share": "1.95%"},
            {"Team": 4, "Management Index": 0, "Agents": 2, "Marketing Investment": 0, "Product Quality Index": 0, "Price": 24444, "Sales Volume": 0, "Market Share": "0.00%"},
            {"Team": 5, "Management Index": 2346.02, "Agents": 1, "Marketing Investment": 388, "Product Quality Index": 6.97, "Price": 24388, "Sales Volume": 1324, "Market Share": "1.88%"},
            {"Team": 6, "Management Index": 11.45, "Agents": 1, "Marketing Investment": 8888, "Product Quality Index": 730.00, "Price": 24200, "Sales Volume": 2802, "Market Share": "3.98%"},
            {"Team": 7, "Management Index": 0, "Agents": 1, "Marketing Investment": 0, "Product Quality Index": 0, "Price": 24444, "Sales Volume": 0, "Market Share": "0.00%"},
            {"Team": 8, "Management Index": 2503.12, "Agents": 1, "Marketing Investment": 0, "Product Quality Index": 0, "Price": 24999, "Sales Volume": 42, "Market Share": "0.05%"},
            {"Team": 9, "Management Index": 2.01, "Agents": 2, "Marketing Investment": 3333, "Product Quality Index": 10.00, "Price": 24333, "Sales Volume": 70, "Market Share": "0.09%"},
            {"Team": 10, "Management Index": 0, "Agents": 4, "Marketing Investment": 4777777, "Product Quality Index": 0, "Price": 24377, "Sales Volume": 3752, "Market Share": "5.32%"},
            {"Team": 12, "Management Index": 0, "Agents": 1, "Marketing Investment": 0, "Product Quality Index": 568.92, "Price": 24333, "Sales Volume": 1126, "Market Share": "1.59%"},
        ]
    },
    "r2_market_hangzhou": {
        "market_info": {
            "Population": 2500000,
            "Penetration": "1.43%",
            "Market Size": 35750,
            "Total Sales Volume": 12088,
            "Avg. Price": 14517
        },
        "teams": [
            {"Team": 2, "Management Index": 0, "Agents": 1, "Marketing Investment": 139, "Product Quality Index": 0, "Price": 24688, "Sales Volume": 0, "Market Share": "0.00%"},
            {"Team": 3, "Management Index": 0, "Agents": 1, "Marketing Investment": 500, "Product Quality Index": 0.62, "Price": 24444, "Sales Volume": 21, "Market Share": "0.05%"},
            {"Team": 5, "Management Index": 2346.02, "Agents": 1, "Marketing Investment": 388, "Product Quality Index": 6.97, "Price": 24388, "Sales Volume": 1323, "Market Share": "3.70%"},
            {"Team": 6, "Management Index": 11.45, "Agents": 1, "Marketing Investment": 8888, "Product Quality Index": 730.00, "Price": 24200, "Sales Volume": 2802, "Market Share": "7.83%"},
            {"Team": 8, "Management Index": 2503.12, "Agents": 1, "Marketing Investment": 0, "Product Quality Index": 0, "Price": 22999, "Sales Volume": 3866, "Market Share": "10.81%"},
            {"Team": 9, "Management Index": 2.01, "Agents": 2, "Marketing Investment": 3333, "Product Quality Index": 10.00, "Price": 24333, "Sales Volume": 32, "Market Share": "0.08%"},
            {"Team": 10, "Management Index": 0, "Agents": 2, "Marketing Investment": 2763606, "Product Quality Index": 0, "Price": 24377, "Sales Volume": 3748, "Market Share": "10.48%"},
            {"Team": 12, "Management Index": 0, "Agents": 1, "Marketing Investment": 0, "Product Quality Index": 568.92, "Price": 24333, "Sales Volume": 296, "Market Share": "0.82%"},
        ]
    },
    "r2_sales": {
        "agents": [
            {"Market": "Shanghai", "Previous": 1, "Change": "+3", "After": 4, "Change Cost": 900000, "Marketing Investment": 6000000},
            {"Market": "Chengdu", "Previous": 1, "Change": "+1", "After": 2, "Change Cost": 300000, "Marketing Investment": 3333},
            {"Market": "Hangzhou", "Previous": 1, "Change": "+1", "After": 2, "Change Cost": 300000, "Marketing Investment": 3333},
        ],
        "market_sales": [
            {"Market": "Shanghai", "Competitive Power": "15.88%", "Sales Volume": 7778, "Market Share": "5.89%", "Price": 24333, "Sales": 189262074},
            {"Market": "Chengdu", "Competitive Power": "0.09%", "Sales Volume": 70, "Market Share": "0.09%", "Price": 24333, "Sales": 1703310},
            {"Market": "Hangzhou", "Competitive Power": "0.08%", "Sales Volume": 32, "Market Share": "0.08%", "Price": 24333, "Sales": 778656},
        ]
    },

    # r1 folder - final round
    "r1_market_shanghai": {
        "market_info": {
            "Population": 6000000,
            "Penetration": "2.00%",
            "Market Size": 120000,
            "Total Sales Volume": 8179,
            "Avg. Price": 10770
        },
        "teams": [
            {"Team": 2, "Management Index": 2699.17, "Agents": 1, "Marketing Investment": 1213, "Product Quality Index": 2.79, "Price": 24688, "Sales Volume": 412, "Market Share": "0.34%"},
            {"Team": 3, "Management Index": 0, "Agents": 1, "Marketing Investment": 500, "Product Quality Index": 0.06, "Price": 24444, "Sales Volume": 72, "Market Share": "0.06%"},
            {"Team": 5, "Management Index": 0, "Agents": 1, "Marketing Investment": 3333, "Product Quality Index": 629.94, "Price": 24888, "Sales Volume": 533, "Market Share": "0.44%"},
            {"Team": 6, "Management Index": 1586.71, "Agents": 1, "Marketing Investment": 0, "Product Quality Index": 1.00, "Price": 24440, "Sales Volume": 1398, "Market Share": "1.16%"},
            {"Team": 7, "Management Index": 1395.80, "Agents": 1, "Marketing Investment": 0, "Product Quality Index": 0, "Price": 24444, "Sales Volume": 1765, "Market Share": "1.47%"},
            {"Team": 8, "Management Index": 60.38, "Agents": 1, "Marketing Investment": 200000, "Product Quality Index": 492.12, "Price": 23999, "Sales Volume": 678, "Market Share": "0.56%"},
            {"Team": 9, "Management Index": 1105.07, "Agents": 1, "Marketing Investment": 3333, "Product Quality Index": 1.01, "Price": 24443, "Sales Volume": 1251, "Market Share": "1.04%"},
            {"Team": 11, "Management Index": 10.46, "Agents": 1, "Marketing Investment": 20000, "Product Quality Index": 12.65, "Price": 23750, "Sales Volume": 166, "Market Share": "0.13%"},
            {"Team": 12, "Management Index": 846.32, "Agents": 1, "Marketing Investment": 0, "Product Quality Index": 0, "Price": 24488, "Sales Volume": 1904, "Market Share": "1.58%"},
        ]
    },
    "r1_market_chengdu": {
        "market_info": {
            "Population": 4000000,
            "Penetration": "1.60%",
            "Market Size": 64000,
            "Total Sales Volume": 7821,
            "Avg. Price": 10590
        },
        "teams": [
            {"Team": 4, "Management Index": 1582.29, "Agents": 2, "Marketing Investment": 0, "Product Quality Index": 0, "Price": 24444, "Sales Volume": 2561, "Market Share": "4.00%"},
            {"Team": 5, "Management Index": 0, "Agents": 1, "Marketing Investment": 3333, "Product Quality Index": 629.94, "Price": 24888, "Sales Volume": 277, "Market Share": "0.43%"},
            {"Team": 6, "Management Index": 1586.71, "Agents": 1, "Marketing Investment": 0, "Product Quality Index": 1.00, "Price": 24440, "Sales Volume": 1158, "Market Share": "1.80%"},
            {"Team": 7, "Management Index": 1395.80, "Agents": 1, "Marketing Investment": 0, "Product Quality Index": 0, "Price": 24444, "Sales Volume": 825, "Market Share": "1.28%"},
            {"Team": 8, "Management Index": 60.38, "Agents": 1, "Marketing Investment": 150000, "Product Quality Index": 492.12, "Price": 21999, "Sales Volume": 678, "Market Share": "1.05%"},
            {"Team": 9, "Management Index": 1105.07, "Agents": 1, "Marketing Investment": 3333, "Product Quality Index": 1.01, "Price": 24443, "Sales Volume": 713, "Market Share": "1.11%"},
            {"Team": 10, "Management Index": 2430.58, "Agents": 1, "Marketing Investment": 0, "Product Quality Index": 0, "Price": 24177, "Sales Volume": 1222, "Market Share": "1.90%"},
            {"Team": 12, "Management Index": 846.32, "Agents": 1, "Marketing Investment": 0, "Product Quality Index": 0, "Price": 24488, "Sales Volume": 387, "Market Share": "0.60%"},
        ]
    },
    "r1_market_hangzhou": {
        "market_info": {
            "Population": 2500000,
            "Penetration": "1.30%",
            "Market Size": 32500,
            "Total Sales Volume": 5504,
            "Avg. Price": 10529
        },
        "teams": [
            {"Team": 2, "Management Index": 2699.17, "Agents": 1, "Marketing Investment": 1213, "Product Quality Index": 2.79, "Price": 24688, "Sales Volume": 411, "Market Share": "1.26%"},
            {"Team": 3, "Management Index": 0, "Agents": 1, "Marketing Investment": 1100000, "Product Quality Index": 0.06, "Price": 24444, "Sales Volume": 2573, "Market Share": "7.91%"},
            {"Team": 8, "Management Index": 60.38, "Agents": 1, "Marketing Investment": 100000, "Product Quality Index": 492.12, "Price": 19999, "Sales Volume": 676, "Market Share": "2.08%"},
            {"Team": 9, "Management Index": 1105.07, "Agents": 1, "Marketing Investment": 3333, "Product Quality Index": 1.01, "Price": 24443, "Sales Volume": 622, "Market Share": "1.91%"},
            {"Team": 10, "Management Index": 2430.58, "Agents": 1, "Marketing Investment": 0, "Product Quality Index": 0, "Price": 24177, "Sales Volume": 1222, "Market Share": "3.76%"},
        ]
    },
    "r1_sales": {
        "agents": [
            {"Market": "Shanghai", "Previous": 0, "Change": "+1", "After": 1, "Change Cost": 300000, "Marketing Investment": 3333},
            {"Market": "Chengdu", "Previous": 0, "Change": "+1", "After": 1, "Change Cost": 300000, "Marketing Investment": 3333},
            {"Market": "Hangzhou", "Previous": 0, "Change": "+1", "After": 1, "Change Cost": 300000, "Marketing Investment": 3333},
        ],
        "market_sales": [
            {"Market": "Shanghai", "Competitive Power": "0.88%", "Sales Volume": 1251, "Market Share": "1.04%", "Price": 24443, "Sales": 30578193},
            {"Market": "Chengdu", "Competitive Power": "0.80%", "Sales Volume": 713, "Market Share": "1.11%", "Price": 24443, "Sales": 17427859},
            {"Market": "Hangzhou", "Competitive Power": "1.32%", "Sales Volume": 622, "Market Share": "1.91%", "Price": 24443, "Sales": 15203546},
        ]
    },
}

def save_market_to_excel(round_name, market_name, market_data, base_dir):
    """Save market data to Excel file"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Market Report"

    # Market Info
    ws.append(["Market Report - " + market_name])
    ws.append([])
    ws.append(["Population", "Penetration", "Market Size", "Total Sales Volume", "Avg. Price"])
    mi = market_data["market_info"]
    ws.append([mi["Population"], mi["Penetration"], mi["Market Size"], mi["Total Sales Volume"], mi["Avg. Price"]])
    ws.append([])

    # Teams data
    ws.append(["Team", "Management Index", "Agents", "Marketing Investment", "Product Quality Index", "Price", "Sales Volume", "Market Share"])
    for team in market_data["teams"]:
        ws.append([
            team["Team"],
            team["Management Index"],
            team["Agents"],
            team["Marketing Investment"],
            team["Product Quality Index"],
            team["Price"],
            team["Sales Volume"],
            team["Market Share"]
        ])

    # Save
    filename = os.path.join(base_dir, round_name, f"{round_name}_market_{market_name.lower()}.xlsx")
    os.makedirs(os.path.dirname(filename), exist_ok=True)
    wb.save(filename)
    print(f"Saved: {filename}")

def save_sales_to_excel(round_name, sales_data, base_dir):
    """Save sales data to Excel file"""
    wb = Workbook()

    # Agents sheet
    ws1 = wb.active
    ws1.title = "Agents"
    ws1.append(["Sales - Agents"])
    ws1.append([])
    ws1.append(["Market", "Previous", "Change", "After", "Change Cost", "Marketing Investment"])
    for agent in sales_data["agents"]:
        ws1.append([
            agent["Market"],
            agent["Previous"],
            agent["Change"],
            agent["After"],
            agent["Change Cost"],
            agent["Marketing Investment"]
        ])

    # Market Sales sheet
    ws2 = wb.create_sheet("Market Sales")
    ws2.append(["Market", "Competitive Power", "Sales Volume", "Market Share", "Price", "Sales"])
    for ms in sales_data["market_sales"]:
        ws2.append([
            ms["Market"],
            ms["Competitive Power"],
            ms["Sales Volume"],
            ms["Market Share"],
            ms["Price"],
            ms["Sales"]
        ])

    # Save
    filename = os.path.join(base_dir, round_name, f"{round_name}_sales.xlsx")
    os.makedirs(os.path.dirname(filename), exist_ok=True)
    wb.save(filename)
    print(f"Saved: {filename}")

def save_summary_to_excel(base_dir):
    """Save a summary of Team 9's performance"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Team 9 Summary"

    ws.append(["Round", "Market", "Agents", "Market Share", "Sales Volume", "Sales (RMB)", "Rank"])
    ws.append([])

    # r4 - best round
    ws.append(["r4", "Shanghai", 10, "15.74%", 25134, 530603874, "1st"])
    ws.append(["r4", "Chengdu", 8, "19.23%", 16312, 344362632, "1st"])
    ws.append(["r4", "Hangzhou", 8, "26.86%", 11554, 256752988, "1st"])
    ws.append([])

    # r3
    ws.append(["r3", "Shanghai", 7, "7.08%", 10292, 228708824, "-"])
    ws.append(["r3", "Chengdu", 5, "9.14%", 7060, 156887320, "-"])
    ws.append(["r3", "Hangzhou", 5, "16.93%", 6648, 146249352, "-"])
    ws.append([])

    # r2
    ws.append(["r2", "Shanghai", 4, "5.89%", 7778, 189262074, "-"])
    ws.append(["r2", "Chengdu", 2, "0.09%", 70, 1703310, "-"])
    ws.append(["r2", "Hangzhou", 2, "0.08%", 32, 778656, "-"])
    ws.append([])

    # r1
    ws.append(["r1", "Shanghai", 1, "1.04%", 1251, 30578193, "-"])
    ws.append(["r1", "Chengdu", 1, "1.11%", 713, 17427859, "-"])
    ws.append(["r1", "Hangzhou", 1, "1.91%", 622, 15203546, "-"])

    filename = os.path.join(base_dir, "team9_summary.xlsx")
    wb.save(filename)
    print(f"Saved: {filename}")

def main():
    base_dir = "/mnt/c/Users/david/documents/ASDAN/表格/obos"

    # Save all market reports
    for round_name in ["r1", "r2", "r3", "r4"]:
        for market in ["Shanghai", "Chengdu", "Hangzhou"]:
            key = f"{round_name}_market_{market.lower()}"
            if key in data:
                save_market_to_excel(round_name, market, data[key], base_dir)

        # Save sales data
        sales_key = f"{round_name}_sales"
        if sales_key in data:
            save_sales_to_excel(round_name, data[sales_key], base_dir)

    # Save summary
    save_summary_to_excel(base_dir)
    print("\nAll Excel files created successfully!")

if __name__ == "__main__":
    main()
