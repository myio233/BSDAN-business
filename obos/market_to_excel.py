#!/usr/bin/env python3
import os
import re
import time
from PIL import Image
import pytesseract
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def smart_split(line):
    """Split line by whitespace but keep currency symbols with numbers"""
    if not line or line.strip() == '':
        return None

    # First, replace currency symbols followed by space with currency+number
    line = re.sub(r'([¥$€])\s+', r'\1', line)

    # Now split by whitespace
    parts = line.strip().split()
    return parts if parts else None


def process_folder(folder_path):
    """Process market report in folder"""
    print(f"\nProcessing: {os.path.basename(folder_path)}")

    market_img = os.path.join(folder_path, "market_report.png")
    if not os.path.exists(market_img):
        print("  No market_report.png")
        return

    # Get OCR text
    img = Image.open(market_img)
    text = pytesseract.image_to_string(img)
    lines = text.split('\n')

    # Find city sections
    cities = []
    current_city = None
    start_idx = 0

    for i, line in enumerate(lines):
        if "Market Report - " in line:
            if current_city:
                cities.append((current_city, start_idx, i))
            current_city = line.replace("Market Report - ", "").strip()
            start_idx = i

    if current_city:
        cities.append((current_city, start_idx, len(lines)))

    if not cities:
        cities = [("Market", 0, len(lines))]

    print(f"  Found {len(cities)} cities")

    # Process each city
    for city_name, start, end in cities:
        wb = Workbook()
        ws = wb.active
        ws.title = "Market Report"

        # Styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        city_lines = lines[start:end]
        rows_written = 0

        for line in city_lines:
            parts = smart_split(line)
            if not parts:
                continue

            # Skip garbage lines with too many special chars
            garbage_count = sum(1 for c in ''.join(parts) if c in '|#@€[]{}')
            if garbage_count > 3:
                continue

            # Write row
            ws.append(parts)
            rows_written += 1

            # Apply header style
            if "Population" in line and "Penetration" in line:
                for cell in ws[ws.max_row]:
                    cell.fill = header_fill
                    cell.font = header_font
            elif "Team" in line and ("Management" in line or "Index" in line):
                for cell in ws[ws.max_row]:
                    cell.fill = header_fill
                    cell.font = header_font

        # Adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        # Save file
        safe_name = re.sub(r'[^\w\s-]', '', city_name).replace(' ', '_')
        folder_name = os.path.basename(folder_path)
        filename = f"{folder_name}_market_{safe_name}.xlsx"
        filepath = os.path.join(folder_path, filename)

        # Try to save
        try:
            if os.path.exists(filepath):
                try:
                    os.remove(filepath)
                    time.sleep(0.1)
                except:
                    filename = f"{folder_name}_market_{safe_name}_new.xlsx"
                    filepath = os.path.join(folder_path, filename)

            wb.save(filepath)
            print(f"    Saved: {filename} ({rows_written} rows)")
        except Exception as e:
            print(f"    Error saving {city_name}: {e}")


def main():
    base_dir = "/mnt/c/Users/david/documents/ASDAN/表格/WYEF"
    folders = ["r-1", "r1", "r2", "r3", "r4", "r5", "r6", "r7"]

    for f in folders:
        folder_path = os.path.join(base_dir, f)
        if os.path.exists(folder_path):
            process_folder(folder_path)


if __name__ == "__main__":
    main()
