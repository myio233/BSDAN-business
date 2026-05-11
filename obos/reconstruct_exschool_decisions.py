#!/usr/bin/env python3
from __future__ import annotations

import math
import sys
import os
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
SCRIPT_DIR = Path(__file__).resolve().parent
EXSCHOOL_DIR = ROOT / "exschool"
MARKET_REPORT_DIR = Path(os.environ.get("EXSCHOOL_MARKET_REPORT_DIR", str(EXSCHOOL_DIR)))
OUTPUT_DIR = ROOT / "outputs" / "exschool_inferred_decisions"
DESKTOP_TEMPLATE = Path(os.environ.get("EXSCHOOL_DECISION_TEMPLATE", "/mnt/c/Users/david/Desktop/商赛表格.xlsx"))

for import_root in (ROOT, SCRIPT_DIR):
    if str(import_root) not in sys.path:
        sys.path.insert(0, str(import_root))

os.environ.setdefault("EXSCHOOL_SKIP_FIXED_DECISION_BOOTSTRAP", "1")

from analyze_team24_competitiveness import parse_numeric, round_sort_key
from exschool_game.engine import ExschoolSimulator

ROUND_FILE_MAP = {
    "r1": "report4_market_reports.xlsx",
    "r2": "report3_market_reports.xlsx",
    "r3": "report2_market_reports_fixed.xlsx",
    "r4": "report1_market_reports_fixed.xlsx",
}
ALL_MARKETS = ["Shanghai", "Chengdu", "Wuhan", "Wuxi", "Ningbo"]
CITY_LABELS = {
    "Shanghai": "上海",
    "Chengdu": "成都",
    "Wuhan": "武汉",
    "Wuxi": "无锡",
    "Ningbo": "宁波",
}
SOURCE_DIRECT = "market_report"
SOURCE_ESTIMATED = "estimated"
SOURCE_ROUND_CONTEXT = "round_context"
SOURCE_SIMULATOR_ACTUAL = "simulator_actual"


@dataclass
class RoundProxy:
    round_id: str
    loan_limit: float
    interest_rate: float
    worker_salary: float
    engineer_salary: float
    components_productivity: float
    products_productivity: float
    component_material_price: float
    product_material_price: float
    component_storage_unit: float
    product_storage_unit: float


ROUND_SUMMARY_PROVENANCE_DEFAULTS = {
    "starting_cash_est_provenance": SOURCE_ESTIMATED,
    "starting_debt_est_provenance": SOURCE_ESTIMATED,
    "loan_delta_est_provenance": SOURCE_ESTIMATED,
    "products_planned_est_provenance": SOURCE_ESTIMATED,
    "quality_investment_est_provenance": SOURCE_ESTIMATED,
    "workers_est_provenance": SOURCE_ESTIMATED,
    "engineers_est_provenance": SOURCE_ESTIMATED,
    "worker_salary_est_provenance": SOURCE_ROUND_CONTEXT,
    "engineer_salary_est_provenance": SOURCE_ROUND_CONTEXT,
    "management_investment_est_provenance": SOURCE_ESTIMATED,
    "component_storage_unit_est_provenance": SOURCE_ROUND_CONTEXT,
    "product_storage_unit_est_provenance": SOURCE_ROUND_CONTEXT,
    "market_report_cost_est_provenance": SOURCE_ESTIMATED,
    "ending_cash_est_provenance": SOURCE_ESTIMATED,
    "ending_debt_est_provenance": SOURCE_ESTIMATED,
}


def currency(value: float) -> str:
    return f"¥{value:,.0f}"


def parse_market_reports() -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for round_id, file_name in ROUND_FILE_MAP.items():
        path = MARKET_REPORT_DIR / file_name
        xl = pd.ExcelFile(path)
        for market in xl.sheet_names:
            df = pd.read_excel(path, sheet_name=market, header=None)
            header_idx = None
            summary_idx = None
            for idx in range(len(df)):
                first = str(df.iloc[idx, 0]).strip() if pd.notna(df.iloc[idx, 0]) else ""
                if first == "Population":
                    summary_idx = idx
                if first == "Team":
                    header_idx = idx
                    break
            if summary_idx is None or header_idx is None:
                continue

            summary_row = summary_idx + 1
            population = parse_numeric(df.iloc[summary_row, 0])
            penetration = parse_numeric(df.iloc[summary_row, 1], percent=True)
            market_size = parse_numeric(df.iloc[summary_row, 2])
            total_sales_volume = parse_numeric(df.iloc[summary_row, 3])
            avg_price = parse_numeric(df.iloc[summary_row, 4])

            row_idx = header_idx + 1
            while row_idx < len(df):
                team = df.iloc[row_idx, 0]
                if pd.isna(team) or not str(team).strip().isdigit():
                    break
                team = str(team).strip()
                agents_after = int(parse_numeric(df.iloc[row_idx, 2]) or 0)
                marketing_investment = float(parse_numeric(df.iloc[row_idx, 3]) or 0.0)
                rows.append(
                    {
                        "round_id": round_id,
                        "market": market,
                        "team": team,
                        "population": float(population or 0.0),
                        "penetration": float(penetration or 0.0),
                        "market_size": float(market_size or 0.0),
                        "total_sales_volume": float(total_sales_volume or 0.0),
                        "avg_price": float(avg_price or 0.0),
                        "management_index": float(parse_numeric(df.iloc[row_idx, 1]) or 0.0),
                        "agents_after": agents_after,
                        "marketing_investment": marketing_investment,
                        "quality_index": float(parse_numeric(df.iloc[row_idx, 4]) or 0.0),
                        "price": float(parse_numeric(df.iloc[row_idx, 5]) or 0.0),
                        "sales_volume": float(parse_numeric(df.iloc[row_idx, 6]) or 0.0),
                        "market_share": float(parse_numeric(df.iloc[row_idx, 7], percent=True) or 0.0),
                        "selected": 1,
                        "file_name": file_name,
                        "market_index": marketing_investment * (1 + 0.1 * agents_after),
                    }
                )
                row_idx += 1
    return pd.DataFrame(rows)


def build_round_proxies(simulator: ExschoolSimulator) -> dict[str, RoundProxy]:
    proxies: dict[str, RoundProxy] = {}
    for round_id in ["r1", "r2", "r3", "r4"]:
        ctx = simulator.round_contexts[round_id]
        proxies[round_id] = RoundProxy(
            round_id=round_id,
            loan_limit=float(ctx["loan_limit"]),
            interest_rate=float(ctx["interest_rate"]),
            worker_salary=float(ctx["worker_salary_actual"]),
            engineer_salary=float(ctx["engineer_salary_actual"]),
            components_productivity=float(ctx["components_productivity"]),
            products_productivity=float(ctx["products_productivity"]),
            component_material_price=float(ctx["component_material_price"]),
            product_material_price=float(ctx["product_material_price"]),
            component_storage_unit=float(ctx["component_storage_unit_cost"]),
            product_storage_unit=float(ctx["product_storage_unit_cost"]),
        )
    return proxies


def ensure_full_grid(decisions: pd.DataFrame) -> pd.DataFrame:
    rounds = ["r1", "r2", "r3", "r4"]
    teams = sorted(decisions["team"].dropna().unique(), key=int)
    idx = pd.MultiIndex.from_product([teams, rounds, ALL_MARKETS], names=["team", "round_id", "market"])
    full = decisions.set_index(["team", "round_id", "market"]).reindex(idx).reset_index()

    for col in ["population", "penetration", "market_size", "total_sales_volume", "avg_price"]:
        full[col] = full.groupby(["round_id", "market"])[col].transform("max")

    numeric_fill = {
        "management_index": 0.0,
        "agents_after": 0.0,
        "marketing_investment": 0.0,
        "quality_index": 0.0,
        "price": 0.0,
        "sales_volume": 0.0,
        "market_share": 0.0,
        "selected": 0.0,
        "market_index": 0.0,
    }
    for col, default in numeric_fill.items():
        full[col] = full[col].fillna(default)

    full["agents_after"] = full["agents_after"].astype(int)
    full["selected"] = full["selected"].astype(int)
    open_pairs = set(map(tuple, full[full["selected"].eq(1)][["round_id", "market"]].drop_duplicates().to_records(index=False)))
    full["market_open"] = [1 if (round_id, market) in open_pairs else 0 for round_id, market in zip(full["round_id"], full["market"])]
    full["selected"] = [sel if open_flag else 0 for sel, open_flag in zip(full["selected"], full["market_open"])]
    return full


def add_agent_deltas(full: pd.DataFrame) -> pd.DataFrame:
    out = full.sort_values(["team", "market", "round_id"], key=lambda s: s.map(round_sort_key) if s.name == "round_id" else s).copy()
    out["agents_before"] = out.groupby(["team", "market"])["agents_after"].shift(1).fillna(0).astype(int)
    out["agent_change"] = (out["agents_after"] - out["agents_before"]).astype(int)
    return out


def override_team13_markets(full: pd.DataFrame, simulator: ExschoolSimulator) -> pd.DataFrame:
    out = full.copy()
    team = "13"
    for round_id in ["r1", "r2", "r3", "r4"]:
        ctx = simulator.round_contexts[round_id]
        total_people = int(ctx["workers_actual"]) + int(ctx["engineers_actual"])
        management_index = (
            float(ctx["management_investment_actual"]) / total_people
            if float(ctx["management_investment_actual"]) > 0 and total_people > 0
            else float(
                out[
                    (out["team"] == team)
                    & (out["round_id"] == round_id)
                    & (out["selected"] == 1)
                ]["management_index"].replace(0, pd.NA).dropna().median()
                or 0.0
            )
        )
        quality_index = (
            float(ctx["quality_investment_actual"]) / float(ctx["products_produced_actual"])
            if float(ctx["products_produced_actual"]) > 0
            else 0.0
        )
        for market in ALL_MARKETS:
            defaults = ctx["market_defaults"][market]
            mask = (out["team"] == team) & (out["round_id"] == round_id) & (out["market"] == market)
            selected = int(
                int(defaults.get("actual_after", 0) or 0) > 0
                or float(defaults.get("actual_marketing_investment", 0.0) or 0.0) > 0
                or float(defaults.get("actual_sales_volume", 0.0) or 0.0) > 0
            )
            out.loc[mask, "selected"] = selected
            out.loc[mask, "management_index"] = management_index if selected else 0.0
            out.loc[mask, "quality_index"] = quality_index if selected else 0.0
            out.loc[mask, "agents_after"] = int(defaults.get("actual_after", 0) or 0)
            out.loc[mask, "marketing_investment"] = float(defaults.get("actual_marketing_investment", 0.0) or 0.0)
            out.loc[mask, "price"] = float(defaults.get("actual_price", 0.0) or 0.0) if selected else 0.0
            out.loc[mask, "sales_volume"] = float(defaults.get("actual_sales_volume", 0.0) or 0.0)
            out.loc[mask, "market_share"] = float(defaults.get("actual_market_share", 0.0) or 0.0)
            out.loc[mask, "market_index"] = float(out.loc[mask, "marketing_investment"].iloc[0]) * (
                1 + 0.1 * int(out.loc[mask, "agents_after"].iloc[0])
            )
    return out


def round_summary_frame(full: pd.DataFrame, proxies: dict[str, RoundProxy]) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    teams = sorted(full["team"].dropna().unique(), key=int)

    starting_cash: dict[str, float] = {team: 15_000_000.0 for team in teams}
    starting_debt: dict[str, float] = {team: 0.0 for team in teams}

    for team in teams:
        team_df = full[full["team"] == team].copy()
        for round_id in ["r1", "r2", "r3", "r4"]:
            proxy = proxies[round_id]
            round_df = team_df[team_df["round_id"] == round_id].copy()
            selected_df = round_df[round_df["selected"] == 1].copy()

            sales_units = float(selected_df["sales_volume"].sum())
            sales_revenue = float((selected_df["sales_volume"] * selected_df["price"]).sum())
            products_planned = int(math.ceil(sales_units))

            engineers = int(math.ceil(products_planned / max(proxy.products_productivity, 1.0))) if products_planned > 0 else 0
            workers = int(math.ceil((products_planned * 7) / max(proxy.components_productivity, 1.0))) if products_planned > 0 else 0

            management_series = selected_df["management_index"].replace(0, pd.NA).dropna()
            quality_series = selected_df["quality_index"].replace(0, pd.NA).dropna()
            management_index = float(management_series.median()) if not management_series.empty else 0.0
            quality_index = float(quality_series.median()) if not quality_series.empty else 0.0
            management_investment = int(round(management_index * (workers + engineers)))
            quality_investment = int(round(quality_index * products_planned))

            worker_salary = int(round(proxy.worker_salary))
            engineer_salary = int(round(proxy.engineer_salary))

            component_units = products_planned * 7
            workers_salary_cost = workers * worker_salary * 3
            engineers_salary_cost = engineers * engineer_salary * 3
            component_material_cost = int(round(component_units * proxy.component_material_price))
            product_material_cost = int(round(products_planned * proxy.product_material_price))
            component_storage_cost = int(round(component_units * proxy.component_storage_unit))
            product_storage_cost = int(round(products_planned * proxy.product_storage_unit))
            agent_change_cost = int(
                sum((300_000 if change >= 0 else 100_000) * abs(change) for change in selected_df["agent_change"].tolist())
            )
            marketing_investment_total = int(round(selected_df["marketing_investment"].sum()))
            market_report_cost = int(round(200_000 * len(selected_df)))

            pre_revenue_cost = (
                workers_salary_cost
                + engineers_salary_cost
                + component_material_cost
                + product_material_cost
                + component_storage_cost
                + product_storage_cost
                + agent_change_cost
                + marketing_investment_total
                + quality_investment
                + management_investment
                + market_report_cost
            )
            loan_required = max(0.0, pre_revenue_cost - starting_cash[team])
            loan_delta = min(proxy.loan_limit, math.ceil(loan_required))
            principal_after = max(starting_debt[team] + loan_delta, 0.0)
            interest = int(round(principal_after * proxy.interest_rate))
            pretax_profit = sales_revenue - pre_revenue_cost - interest
            tax = int(round(max(pretax_profit, 0.0) * 0.25))
            ending_cash = starting_cash[team] + loan_delta - pre_revenue_cost + sales_revenue - interest - tax
            ending_debt = principal_after + interest

            rows.append(
                {
                    "team": team,
                    "round_id": round_id,
                    "starting_cash_est": int(round(starting_cash[team])),
                    "starting_debt_est": int(round(starting_debt[team])),
                    "loan_delta_est": int(round(loan_delta)),
                    "products_planned_est": products_planned,
                    "quality_investment_est": quality_investment,
                    "workers_est": workers,
                    "engineers_est": engineers,
                    "worker_salary_est": worker_salary,
                    "engineer_salary_est": engineer_salary,
                    "management_investment_est": management_investment,
                    "component_storage_unit_est": int(round(proxy.component_storage_unit)),
                    "product_storage_unit_est": int(round(proxy.product_storage_unit)),
                    "management_index_source": management_index,
                    "quality_index_source": quality_index,
                    "sales_units_source": int(round(sales_units)),
                    "sales_revenue_source": int(round(sales_revenue)),
                    "market_report_cost_est": market_report_cost,
                    "ending_cash_est": int(round(ending_cash)),
                    "ending_debt_est": int(round(ending_debt)),
                    **ROUND_SUMMARY_PROVENANCE_DEFAULTS,
                }
            )

            starting_cash[team] = ending_cash
            starting_debt[team] = ending_debt

    return pd.DataFrame(rows)


def build_numeric_export(full: pd.DataFrame, rounds: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for (team, round_id), round_df in full.groupby(["team", "round_id"], sort=False):
        round_meta = rounds[(rounds["team"] == team) & (rounds["round_id"] == round_id)].iloc[0].to_dict()
        row: dict[str, Any] = {
            "team": team,
            "round_id": round_id,
            "loan_delta": int(round_meta["loan_delta_est"]),
            "loan_delta_provenance": round_meta["loan_delta_est_provenance"],
            "products_planned": int(round_meta["products_planned_est"]),
            "products_planned_provenance": round_meta["products_planned_est_provenance"],
            "quality_investment": int(round_meta["quality_investment_est"]),
            "quality_investment_provenance": round_meta["quality_investment_est_provenance"],
            "workers": int(round_meta["workers_est"]),
            "workers_provenance": round_meta["workers_est_provenance"],
            "engineers": int(round_meta["engineers_est"]),
            "engineers_provenance": round_meta["engineers_est_provenance"],
            "worker_salary": int(round_meta["worker_salary_est"]),
            "worker_salary_provenance": round_meta["worker_salary_est_provenance"],
            "engineer_salary": int(round_meta["engineer_salary_est"]),
            "engineer_salary_provenance": round_meta["engineer_salary_est_provenance"],
            "management_investment": int(round_meta["management_investment_est"]),
            "management_investment_provenance": round_meta["management_investment_est_provenance"],
        }
        for market in ALL_MARKETS:
            market_row = round_df[round_df["market"] == market].iloc[0]
            prefix = market.lower()
            row[f"{prefix}_selected"] = int(market_row["selected"])
            row[f"{prefix}_agents_before"] = int(market_row["agents_before"])
            row[f"{prefix}_agent_change"] = int(market_row["agent_change"])
            row[f"{prefix}_agents_after"] = int(market_row["agents_after"])
            row[f"{prefix}_marketing_investment"] = int(round(float(market_row["marketing_investment"])))
            row[f"{prefix}_price"] = int(round(float(market_row["price"])))
        rows.append(row)
    return pd.DataFrame(rows).sort_values(["team", "round_id"], key=lambda s: s.map(round_sort_key) if s.name == "round_id" else s)


def override_team13_round_summary(round_summary: pd.DataFrame, full: pd.DataFrame, simulator: ExschoolSimulator) -> pd.DataFrame:
    out = round_summary.copy()
    team = "13"
    for round_id in ["r1", "r2", "r3", "r4"]:
        ctx = simulator.round_contexts[round_id]
        total_people = int(ctx["workers_actual"]) + int(ctx["engineers_actual"])
        fallback_management_index = float(
            full[
                (full["team"] == team)
                & (full["round_id"] == round_id)
                & (full["selected"] == 1)
            ]["management_index"].replace(0, pd.NA).dropna().median()
            or 0.0
        )
        management_investment = (
            int(round(float(ctx["management_investment_actual"])))
            if float(ctx["management_investment_actual"]) > 0
            else int(round(fallback_management_index * total_people))
        )
        quality_investment = int(round(float(ctx["quality_investment_actual"])))
        products_planned = int(ctx["products_produced_actual"])
        management_index = management_investment / total_people if total_people > 0 else 0.0
        quality_index = quality_investment / products_planned if products_planned > 0 else 0.0
        mask = (out["team"] == team) & (out["round_id"] == round_id)
        out.loc[mask, "starting_cash_est"] = int(round(float(ctx["starting_cash"])))
        out.loc[mask, "starting_cash_est_provenance"] = SOURCE_SIMULATOR_ACTUAL
        out.loc[mask, "starting_debt_est"] = int(round(float(ctx["starting_debt"])))
        out.loc[mask, "starting_debt_est_provenance"] = SOURCE_SIMULATOR_ACTUAL
        out.loc[mask, "loan_delta_est"] = int(round(float(ctx["actual_loan_delta"])))
        out.loc[mask, "loan_delta_est_provenance"] = SOURCE_SIMULATOR_ACTUAL
        out.loc[mask, "products_planned_est"] = products_planned
        out.loc[mask, "products_planned_est_provenance"] = SOURCE_SIMULATOR_ACTUAL
        out.loc[mask, "quality_investment_est"] = quality_investment
        out.loc[mask, "quality_investment_est_provenance"] = SOURCE_SIMULATOR_ACTUAL
        out.loc[mask, "workers_est"] = int(ctx["workers_actual"])
        out.loc[mask, "workers_est_provenance"] = SOURCE_SIMULATOR_ACTUAL
        out.loc[mask, "engineers_est"] = int(ctx["engineers_actual"])
        out.loc[mask, "engineers_est_provenance"] = SOURCE_SIMULATOR_ACTUAL
        out.loc[mask, "worker_salary_est"] = int(round(float(ctx["worker_salary_actual"])))
        out.loc[mask, "worker_salary_est_provenance"] = SOURCE_ROUND_CONTEXT
        out.loc[mask, "engineer_salary_est"] = int(round(float(ctx["engineer_salary_actual"])))
        out.loc[mask, "engineer_salary_est_provenance"] = SOURCE_ROUND_CONTEXT
        out.loc[mask, "management_investment_est"] = management_investment
        out.loc[mask, "management_investment_est_provenance"] = (
            SOURCE_SIMULATOR_ACTUAL if float(ctx["management_investment_actual"]) > 0 else SOURCE_ESTIMATED
        )
        out.loc[mask, "component_storage_unit_est"] = int(round(float(ctx["component_storage_unit_cost"])))
        out.loc[mask, "component_storage_unit_est_provenance"] = SOURCE_ROUND_CONTEXT
        out.loc[mask, "product_storage_unit_est"] = int(round(float(ctx["product_storage_unit_cost"])))
        out.loc[mask, "product_storage_unit_est_provenance"] = SOURCE_ROUND_CONTEXT
        out.loc[mask, "management_index_source"] = management_index
        out.loc[mask, "quality_index_source"] = quality_index
        out.loc[mask, "sales_units_source"] = int(
            round(
                full[
                    (full["team"] == team)
                    & (full["round_id"] == round_id)
                ]["sales_volume"].sum()
            )
        )
        out.loc[mask, "sales_revenue_source"] = int(
            round(
                (
                    full[
                        (full["team"] == team)
                        & (full["round_id"] == round_id)
                    ]["sales_volume"]
                    * full[
                        (full["team"] == team)
                        & (full["round_id"] == round_id)
                    ]["price"]
                ).sum()
            )
        )
        out.loc[mask, "market_report_cost_est"] = int(round(float(ctx["market_report_cost"])))
        out.loc[mask, "market_report_cost_est_provenance"] = SOURCE_SIMULATOR_ACTUAL
    return out


def autosize(ws) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = max(widths.get(cell.column, 0), len(str(cell.value)))
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = min(max(width + 2, 10), 24)


def init_template_workbook() -> Workbook:
    if DESKTOP_TEMPLATE.exists():
        return load_workbook(DESKTOP_TEMPLATE)
    wb = Workbook()
    wb.active.title = "round1"
    for sheet_name in ["round2", "round3", "round4"]:
        wb.create_sheet(sheet_name)
    return wb


def write_template_workbook(team: str, team_full: pd.DataFrame, team_numeric: pd.DataFrame, output_path: Path) -> None:
    wb = init_template_workbook()
    title_fill = PatternFill("solid", fgColor="1F4E78")
    block_fill = PatternFill("solid", fgColor="D9EAF7")
    market_fill = PatternFill("solid", fgColor="E2F0D9")
    title_font = Font(color="FFFFFF", bold=True, size=14)
    bold = Font(bold=True)

    if "决策汇总" in wb.sheetnames:
        del wb["决策汇总"]
    summary_ws = wb.create_sheet("决策汇总", 0)
    summary_ws.sheet_view.showGridLines = False
    summary_ws["A1"] = f"C{team} 四轮反推决策"
    summary_ws["A1"].fill = title_fill
    summary_ws["A1"].font = title_font
    summary_ws.merge_cells("A1:J1")
    summary_ws["A3"] = "说明"
    summary_ws["B3"] = "市场字段来自图片校验后的市场报表；隐藏字段按统一规则反推。"
    summary_ws["A3"].font = bold

    start_row = 5
    round_sheet_map = {"r1": "round1", "r2": "round2", "r3": "round3", "r4": "round4"}

    for idx, round_id in enumerate(["r1", "r2", "r3", "r4"]):
        row0 = start_row + idx * 12
        num_row = team_numeric[team_numeric["round_id"] == round_id].iloc[0]
        round_full = team_full[team_full["round_id"] == round_id].copy()

        summary_ws[f"A{row0}"] = f"{round_id.upper()} 基础决策"
        summary_ws[f"A{row0}"].fill = block_fill
        summary_ws[f"A{row0}"].font = bold
        summary_ws.merge_cells(start_row=row0, start_column=1, end_row=row0, end_column=9)

        base_items = [
            ("贷款", int(num_row["loan_delta"])),
            ("生产数量", int(num_row["products_planned"])),
            ("质量投资", int(num_row["quality_investment"])),
            ("工人数量", int(num_row["workers"])),
            ("工程师数量", int(num_row["engineers"])),
            ("工人工资", int(num_row["worker_salary"])),
            ("工程师工资", int(num_row["engineer_salary"])),
            ("管理投资", int(num_row["management_investment"])),
        ]
        for offset, (label, value) in enumerate(base_items, start=1):
            summary_ws[f"A{row0 + offset}"] = label
            summary_ws[f"B{row0 + offset}"] = value

        market_header_row = row0 + 1
        for col_idx, title in enumerate(["市场", "是否进入", "期初代理", "代理变化", "期末代理", "市场投资", "价格"], start=4):
            summary_ws.cell(row=market_header_row, column=col_idx, value=title)
            summary_ws.cell(row=market_header_row, column=col_idx).fill = market_fill
            summary_ws.cell(row=market_header_row, column=col_idx).font = bold
        for offset, market in enumerate(ALL_MARKETS, start=1):
            target_row = market_header_row + offset
            market_row = round_full[round_full["market"] == market].iloc[0]
            values = [
                CITY_LABELS[market],
                int(market_row["selected"]),
                int(market_row["agents_before"]),
                int(market_row["agent_change"]),
                int(market_row["agents_after"]),
                int(round(float(market_row["marketing_investment"]))),
                int(round(float(market_row["price"]))),
            ]
            for col_idx, value in enumerate(values, start=4):
                summary_ws.cell(row=target_row, column=col_idx, value=value)

        # Also stamp a decision block onto the original round sheet below the existing template.
        round_ws = wb[round_sheet_map[round_id]]
        base_row = 45
        round_ws[f"A{base_row}"] = f"C{team} {round_id.upper()} 反推原始决策"
        round_ws[f"A{base_row}"].fill = block_fill
        round_ws[f"A{base_row}"].font = bold
        for i, (label, value) in enumerate(base_items, start=1):
            round_ws.cell(row=base_row + i, column=1, value=label)
            round_ws.cell(row=base_row + i, column=2, value=value)
        header_row = base_row + 1
        for col_idx, title in enumerate(["市场", "进入", "期初代理", "代理变化", "期末代理", "市场投资", "价格"], start=4):
            round_ws.cell(row=header_row, column=col_idx, value=title)
            round_ws.cell(row=header_row, column=col_idx).fill = market_fill
            round_ws.cell(row=header_row, column=col_idx).font = bold
        for offset, market in enumerate(ALL_MARKETS, start=1):
            target_row = header_row + offset
            market_row = round_full[round_full["market"] == market].iloc[0]
            values = [
                CITY_LABELS[market],
                int(market_row["selected"]),
                int(market_row["agents_before"]),
                int(market_row["agent_change"]),
                int(market_row["agents_after"]),
                int(round(float(market_row["marketing_investment"]))),
                int(round(float(market_row["price"]))),
            ]
            for col_idx, value in enumerate(values, start=4):
                round_ws.cell(row=target_row, column=col_idx, value=value)

    for ws in wb.worksheets:
        autosize(ws)
    wb.save(output_path)


def write_numeric_workbook(team: str, team_numeric: pd.DataFrame, output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "纯决策数值"
    ws.append(team_numeric.columns.tolist())
    for row in team_numeric.itertuples(index=False):
        ws.append(list(row))
    ws.freeze_panes = "A2"
    ws["A1"].font = Font(bold=True)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    autosize(ws)
    wb.save(output_path)


def write_summary_files(full: pd.DataFrame, round_summary: pd.DataFrame, numeric: pd.DataFrame) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    full.to_excel(OUTPUT_DIR / "all_visible_market_decisions.xlsx", index=False)
    round_summary.to_excel(OUTPUT_DIR / "all_round_reconstruction_summary.xlsx", index=False)
    numeric.to_excel(OUTPUT_DIR / "all_companies_numeric_decisions.xlsx", index=False)


def main() -> None:
    simulator = ExschoolSimulator()
    proxies = build_round_proxies(simulator)

    visible = parse_market_reports()
    full = ensure_full_grid(visible)
    full = add_agent_deltas(full)
    full = override_team13_markets(full, simulator)
    full = add_agent_deltas(full)
    round_summary = round_summary_frame(full, proxies)
    round_summary = override_team13_round_summary(round_summary, full, simulator)
    numeric = build_numeric_export(full, round_summary)

    write_summary_files(full, round_summary, numeric)

    for team in sorted(numeric["team"].unique(), key=int):
        team_full = full[full["team"] == team].copy()
        team_numeric = numeric[numeric["team"] == team].copy()
        write_template_workbook(
            team,
            team_full,
            team_numeric,
            OUTPUT_DIR / f"C{int(team):02d}_反推决策模板.xlsx",
        )
        write_numeric_workbook(
            team,
            team_numeric,
            OUTPUT_DIR / f"C{int(team):02d}_纯决策数值.xlsx",
        )

    assumptions = [
        "显性字段直接来自市场报表：管理指数、质量指数、代理期末数、市场投资、价格、销量、份额。",
        "代理变化 = 本轮代理期末数 - 上轮代理期末数；首轮以上轮为 0。",
        "生产数量按最小可行解处理：等于本轮各市场销量总和，不额外假设库存。",
        "工人/工程师数量按 Team13 当轮产能参数反推的最小可行人数估算。",
        "工人工资/工程师工资采用 Team13 当轮真实工资作为回合同步代理值。",
        "质量投资 = 质量指数 × 生产数量；管理投资 = 管理指数 × (工人 + 工程师)。",
        "贷款取维持售前现金流不为负的最小估算值，并受该轮贷款上限约束。",
        "市场报表无法唯一识别工资、人员、贷款、研发、融资、专利与库存，因此这些字段均为统一假设反推值，不是原始真值。",
    ]
    (OUTPUT_DIR / "ASSUMPTIONS.txt").write_text("\n".join(f"- {line}" for line in assumptions), encoding="utf-8")

    print(f"Output written to: {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
