#!/usr/bin/env python3
import os
import re
import time
from PIL import Image
import pytesseract
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def parse_line_with_positions(img_path):
    """Use OCR with position data to better parse columns"""
    img = Image.open(img_path)

    # Get data with bounding boxes
    data = pytesseract.image_to_data(img, output_type=pytesseract.Output.DICT)

    # Group text by line
    lines = {}
    n_boxes = len(data['text'])

    for i in range(n_boxes):
        if int(data['conf'][i]) < 30:  # Skip low confidence
            continue

        text = data['text'][i].strip()
        if not text:
            continue

        top = data['top'][i]
        left = data['left'][i]
        width = data['width'][i]

        # Group by vertical position (top)
        line_key = round(top / 5) * 5  # Group within 5px
        if line_key not in lines:
            lines[line_key] = []
        lines[line_key].append((left, text, width))

    # Sort lines and their content
    sorted_lines = []
    for line_key in sorted(lines.keys()):
        # Sort by horizontal position
        line_content = sorted(lines[line_key], key=lambda x: x[0])
        texts = [t[1] for t in line_content]

        # Merge currency symbols with numbers
        merged = []
        i = 0
        while i < len(texts):
            t = texts[i]
            if t in ['¥', '$', '€'] and i + 1 < len(texts):
                merged.append(t + texts[i + 1])
                i += 2
            else:
                merged.append(t)
                i += 1
        sorted_lines.append(merged)

    return sorted_lines


def process_market_report(folder_path):
    """Process market report with improved parsing"""
    folder_name = os.path.basename(folder_path)
    print(f"\nProcessing: {folder_name}")

    market_img = os.path.join(folder_path, "market_report.png")
    if not os.path.exists(market_img):
        print("  No market_report.png")
        return

    # Get OCR text first to find cities
    full_text = pytesseract.image_to_string(Image.open(market_img))
    lines = full_text.split('\n')

    # Find city sections
    cities = []
    current_city = None
    city_start = 0

    for i, line in enumerate(lines):
        if "Market Report - " in line:
            if current_city:
                cities.append((current_city, city_start, i))
            current_city = line.replace("Market Report - ", "").strip()
            city_start = i

    if current_city:
        cities.append((current_city, city_start, len(lines)))

    if not cities:
        cities = [("Market", 0, len(lines))]

    print(f"  Found {len(cities)} cities")

    # Get line-by-line with positions
    parsed_lines = parse_line_with_positions(market_img)

    # Process each city
    for city_name, start_line, end_line in cities:
        wb = Workbook()
        ws = wb.active
        ws.title = "Market Report"

        # Styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center")

        rows_written = 0
        city_text = "\n".join(lines[start_line:end_line])

        # Find the city section in parsed lines
        # First, get the approximate range
        approx_start = max(0, start_line - 2)
        approx_end = min(len(parsed_lines), end_line + 2)

        for line_parts in parsed_lines[approx_start:approx_end]:
            if not line_parts:
                continue

            line_str = " ".join(line_parts)

            # Skip if not in city section
            if rows_written == 0 and "Market Report" not in line_str and city_name not in line_str:
                if len(cities) > 1:
                    continue

            # Skip garbage lines
            garbage = sum(1 for c in line_str if c in '|#@[]{}<>')
            if garbage > 2:
                continue

            # Write row
            ws.append(line_parts)
            rows_written += 1

            # Check if header
            is_header = ("Population" in line_str and "Penetration" in line_str) or \
                        ("Team" in line_str and ("Management" in line_str or "Index" in line_str))

            if is_header:
                for cell in ws[ws.max_row]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align

        # Auto-adjust columns
        for column in ws.columns:
            max_len = 0
            col_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 55)

        # Save
        safe_name = re.sub(r'[^\w\s-]', '', city_name).replace(' ', '_')
        filename = f"{folder_name}_market_{safe_name}.xlsx"
        filepath = os.path.join(folder_path, filename)

        # Handle file locks
        for attempt in range(3):
            try:
                if os.path.exists(filepath):
                    os.remove(filepath)
                    time.sleep(0.1)
                wb.save(filepath)
                print(f"    Saved: {filename} ({rows_written} rows)")
                break
            except Exception as e:
                if attempt == 2:
                    filename = f"{folder_name}_market_{safe_name}_v2.xlsx"
                    filepath = os.path.join(folder_path, filename)
                    wb.save(filepath)
                    print(f"    Saved: {filename} ({rows_written} rows)")
                else:
                    time.sleep(0.2)


def main():
    base_dir = "/mnt/c/Users/david/documents/ASDAN/表格/WYEF"
    folders = ["r-1", "r1", "r2", "r3", "r4", "r5", "r6", "r7"]

    for f in folders:
        folder_path = os.path.join(base_dir, f)
        if os.path.isdir(folder_path):
            process_market_report(folder_path)


if __name__ == "__main__":
    main()
