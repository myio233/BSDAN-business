#!/usr/bin/env python3
import os
import re
import time
from azure.ai.formrecognizer import DocumentAnalysisClient
from azure.core.credentials import AzureKeyCredential
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def get_azure_client():
    """Get Azure Document Intelligence client"""
    endpoint = os.environ.get("DOCUMENTINTELLIGENCE_ENDPOINT")
    key = os.environ.get("DOCUMENTINTELLIGENCE_API_KEY")

    # If not in env vars, prompt or use default
    if not endpoint or not key:
        print("Please set DOCUMENTINTELLIGENCE_ENDPOINT and DOCUMENTINTELLIGENCE_API_KEY")
        print("Or enter them below:")
        if not endpoint:
            endpoint = input("Endpoint: ").strip()
        if not key:
            key = input("Key: ").strip()

    if not endpoint or not key:
        raise ValueError("Azure credentials required")

    return DocumentAnalysisClient(
        endpoint=endpoint,
        credential=AzureKeyCredential(key)
    )


def analyze_image_with_azure(image_path):
    """Analyze image using Azure Document Intelligence"""
    client = get_azure_client()

    with open(image_path, "rb") as f:
        poller = client.begin_analyze_document(
            "prebuilt-layout",
            document=f
        )

    result = poller.result()
    return result


def extract_tables_from_result(result):
    """Extract tables from Azure analysis result"""
    all_tables = []

    for table in result.tables:
        table_data = []
        # Create a 2D array for the table
        rows = [["" for _ in range(table.column_count)] for _ in range(table.row_count)]

        for cell in table.cells:
            rows[cell.row_index][cell.column_index] = cell.content

        table_data.extend(rows)
        all_tables.append(table_data)

    # Also extract plain text for city detection
    full_text = ""
    for page in result.pages:
        for line in page.lines:
            full_text += line.content + "\n"

    return all_tables, full_text


def find_city_sections(full_text):
    """Find city sections from text"""
    lines = full_text.split('\n')
    cities = []
    current_city = None
    start_idx = 0

    for i, line in enumerate(lines):
        if "Market Report - " in line:
            if current_city:
                cities.append((current_city, start_idx, i))
            current_city = line.replace("Market Report - ", "").strip()
            start_idx = i

    if current_city:
        cities.append((current_city, start_idx, len(lines)))

    if not cities:
        cities = [("Market", 0, len(lines))]

    return cities


def save_table_to_excel(table_data, output_path, sheet_name="Market Report"):
    """Save table data to Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center")

    for row_idx, row_data in enumerate(table_data):
        ws.append(row_data)

        # Check if this looks like a header row
        row_str = " ".join(str(cell) for cell in row_data)
        is_header = ("Population" in row_str and "Penetration" in row_str) or \
                    ("Team" in row_str and ("Management" in row_str or "Index" in row_str))

        if is_header:
            for cell in ws[ws.max_row]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align

    # Auto-adjust column widths
    for column in ws.columns:
        max_len = 0
        col_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 55)

    # Save
    for attempt in range(3):
        try:
            if os.path.exists(output_path):
                os.remove(output_path)
                time.sleep(0.1)
            wb.save(output_path)
            return True
        except Exception as e:
            if attempt == 2:
                base, ext = os.path.splitext(output_path)
                output_path = f"{base}_v2{ext}"
                wb.save(output_path)
                return True
            time.sleep(0.2)


def process_folder(folder_path):
    """Process a folder with Azure OCR"""
    folder_name = os.path.basename(folder_path)
    print(f"\nProcessing: {folder_name}")

    market_img = os.path.join(folder_path, "market_report.png")
    if not os.path.exists(market_img):
        print("  No market_report.png")
        return

    print("  Analyzing with Azure Document Intelligence...")
    try:
        result = analyze_image_with_azure(market_img)
        tables, full_text = extract_tables_from_result(result)
    except Exception as e:
        print(f"  Azure Error: {e}")
        return

    print(f"  Found {len(tables)} tables")

    # Find city sections
    cities = find_city_sections(full_text)
    print(f"  Found {len(cities)} city sections")

    # If Azure found tables, use them
    if tables:
        # Try to match tables to cities
        for i, table in enumerate(tables):
            city_name = cities[i][0] if i < len(cities) else f"Table_{i+1}"
            safe_name = re.sub(r'[^\w\s-]', '', city_name).replace(' ', '_')
            filename = f"{folder_name}_market_{safe_name}_azure.xlsx"
            filepath = os.path.join(folder_path, filename)

            if save_table_to_excel(table, filepath):
                print(f"    Saved: {filename} ({len(table)} rows)")
    else:
        # Fallback: save the whole thing
        print("  No tables found, saving full text...")
        lines = full_text.split('\n')
        wb = Workbook()
        ws = wb.active
        for line in lines:
            ws.append([line])
        filepath = os.path.join(folder_path, f"{folder_name}_full_text.xlsx")
        wb.save(filepath)
        print(f"    Saved: {folder_name}_full_text.xlsx")


def main():
    base_dir = "/mnt/c/Users/david/documents/ASDAN/表格/WYEF"
    folders = ["r-1", "r1", "r2", "r3", "r4", "r5", "r6", "r7"]

    print("Azure Document Intelligence Market Report Converter")
    print("=" * 50)

    for f in folders:
        folder_path = os.path.join(base_dir, f)
        if os.path.isdir(folder_path):
            process_folder(folder_path)


if __name__ == "__main__":
    main()
