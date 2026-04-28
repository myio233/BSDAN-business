#!/usr/bin/env python3
import os
import re
import time
from PIL import Image as PILImage
from azure.ai.formrecognizer import DocumentAnalysisClient
from azure.core.credentials import AzureKeyCredential
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def get_azure_client():
    """Get Azure Document Intelligence client"""
    endpoint = os.environ.get("DOCUMENTINTELLIGENCE_ENDPOINT")
    key = os.environ.get("DOCUMENTINTELLIGENCE_API_KEY")

    if not endpoint or not key:
        raise ValueError("Azure credentials required")

    return DocumentAnalysisClient(
        endpoint=endpoint,
        credential=AzureKeyCredential(key)
    )


def compress_image_if_needed(image_path, max_size_mb=50):
    """Compress image if it's too large for Azure"""
    size_mb = os.path.getsize(image_path) / (1024 * 1024)

    if size_mb <= max_size_mb:
        return image_path

    print(f"  Compressing image (was {size_mb:.1f}MB)...")
    img = PILImage.open(image_path)

    # Resize if too large
    max_dimension = 10000
    if img.width > max_dimension or img.height > max_dimension:
        ratio = min(max_dimension / img.width, max_dimension / img.height)
        new_size = (int(img.width * ratio), int(img.height * ratio))
        img = img.resize(new_size, PILImage.LANCZOS)

    # Save with compression
    compressed_path = image_path + ".compressed.jpg"
    img.save(compressed_path, "JPEG", quality=85)

    new_size_mb = os.path.getsize(compressed_path) / (1024 * 1024)
    print(f"  Compressed to {new_size_mb:.1f}MB")
    return compressed_path


def analyze_image_with_azure(image_path):
    """Analyze image using Azure Document Intelligence"""
    client = get_azure_client()

    # Compress if needed
    processed_path = compress_image_if_needed(image_path)

    with open(processed_path, "rb") as f:
        poller = client.begin_analyze_document(
            "prebuilt-layout",
            document=f
        )

    result = poller.result()

    # Clean up compressed file
    if processed_path != image_path and os.path.exists(processed_path):
        os.remove(processed_path)

    return result


def merge_related_tables(tables):
    """Merge tables that belong together (header + data)"""
    if len(tables) <= 1:
        return tables

    merged = []
    i = 0
    while i < len(tables):
        table = tables[i]

        # If this table is small (1-3 rows) and next table exists
        if i + 1 < len(tables) and len(table) <= 3:
            next_table = tables[i + 1]

            # Check if column counts match or are compatible
            if len(table[0]) == len(next_table[0]) or abs(len(table[0]) - len(next_table[0])) <= 2:
                # Merge them
                merged_table = table + next_table
                merged.append(merged_table)
                i += 2
                continue

        merged.append(table)
        i += 1

    return merged


def extract_tables_from_result(result):
    """Extract tables from Azure analysis result"""
    all_tables = []

    for table in result.tables:
        table_data = []
        rows = [["" for _ in range(table.column_count)] for _ in range(table.row_count)]

        for cell in table.cells:
            rows[cell.row_index][cell.column_index] = cell.content

        table_data.extend(rows)
        all_tables.append(table_data)

    # Merge related tables
    all_tables = merge_related_tables(all_tables)

    # Extract plain text
    full_text = ""
    for page in result.pages:
        for line in page.lines:
            full_text += line.content + "\n"

    return all_tables, full_text


def find_city_sections(full_text):
    """Find city sections from text"""
    lines = full_text.split('\n')
    cities = []
    current_city = None
    start_idx = 0

    for i, line in enumerate(lines):
        if "Market Report - " in line:
            if current_city:
                cities.append((current_city, start_idx, i))
            current_city = line.replace("Market Report - ", "").strip()
            start_idx = i

    if current_city:
        cities.append((current_city, start_idx, len(lines)))

    if not cities:
        cities = [("Market", 0, len(lines))]

    return cities


def save_table_to_excel(table_data, output_path, sheet_name="Market Report"):
    """Save table data to Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center")

    for row_idx, row_data in enumerate(table_data):
        ws.append(row_data)

        # Check if this looks like a header row
        row_str = " ".join(str(cell) for cell in row_data)
        is_header = ("Population" in row_str and "Penetration" in row_str) or \
                    ("Team" in row_str and ("Management" in row_str or "Index" in row_str))

        if is_header:
            for cell in ws[ws.max_row]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align

    # Auto-adjust column widths
    for column in ws.columns:
        max_len = 0
        col_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 55)

    # Save
    for attempt in range(3):
        try:
            if os.path.exists(output_path):
                os.remove(output_path)
                time.sleep(0.1)
            wb.save(output_path)
            return True
        except Exception as e:
            if attempt == 2:
                base, ext = os.path.splitext(output_path)
                output_path = f"{base}_v3{ext}"
                wb.save(output_path)
                return True
            time.sleep(0.2)


def process_folder(folder_path):
    """Process a folder with Azure OCR"""
    folder_name = os.path.basename(folder_path)
    print(f"\nProcessing: {folder_name}")

    market_img = os.path.join(folder_path, "market_report.png")
    if not os.path.exists(market_img):
        print("  No market_report.png")
        return

    print("  Analyzing with Azure Document Intelligence...")
    try:
        result = analyze_image_with_azure(market_img)
        tables, full_text = extract_tables_from_result(result)
    except Exception as e:
        print(f"  Azure Error: {e}")
        return

    print(f"  Found {len(tables)} tables (after merging)")

    # Find city sections
    cities = find_city_sections(full_text)
    print(f"  Found {len(cities)} city sections")

    # Save tables
    saved_count = 0
    for i, table in enumerate(tables):
        if i < len(cities):
            city_name = cities[i][0]
        else:
            city_name = f"Table_{i+1}"

        safe_name = re.sub(r'[^\w\s-]', '', city_name).replace(' ', '_')
        filename = f"{folder_name}_market_{safe_name}_azure.xlsx"
        filepath = os.path.join(folder_path, filename)

        # Skip very small tables (likely not useful)
        if len(table) < 2:
            continue

        if save_table_to_excel(table, filepath):
            print(f"    Saved: {filename} ({len(table)} rows)")
            saved_count += 1

    print(f"  Total saved: {saved_count} files")


def main():
    base_dir = "/mnt/c/Users/david/documents/ASDAN/表格/WYEF"
    folders = ["r-1", "r1", "r2", "r3", "r4", "r5", "r6", "r7"]

    print("Azure Document Intelligence Market Report Converter (v2)")
    print("=" * 60)

    for f in folders:
        folder_path = os.path.join(base_dir, f)
        if os.path.isdir(folder_path):
            process_folder(folder_path)


if __name__ == "__main__":
    main()
