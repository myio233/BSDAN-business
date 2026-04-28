#!/usr/bin/env python3
from __future__ import annotations

import re
from pathlib import Path

import pandas as pd
import pytesseract
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from PIL import Image


ROOT = Path(__file__).resolve().parents[1]
EXSCHOOL_DIR = ROOT / "exschool"
OUTPUT_DIR = ROOT / "outputs" / "exschool_market_report_exports"

IMAGE_ROUND_MAP = {
    "report_r1.jpg": "r1",
    "report_r2.jpg": "r2",
    "report_r3.jpg": "r3",
    "report_r4.jpg": "r4",
}

ROUND_SOURCE_XLSX = {
    "r1": EXSCHOOL_DIR / "report4_market_reports.xlsx",
    "r2": EXSCHOOL_DIR / "report3_market_reports.xlsx",
    "r3": EXSCHOOL_DIR / "report2_market_reports_fixed.xlsx",
    "r4": EXSCHOOL_DIR / "report1_market_reports_fixed.xlsx",
}


def ocr_image(image_path: Path) -> tuple[str, list[str]]:
    text = pytesseract.image_to_string(Image.open(image_path), lang="eng")
    round_match = re.search(r"Round\s+(\d+)\s+Report", text, re.IGNORECASE)
    round_id = f"r{round_match.group(1)}" if round_match else "unknown"
    cities = re.findall(r"Market Report - ([A-Za-z]+)", text)
    seen: list[str] = []
    for city in cities:
        if city not in seen:
            seen.append(city)
    return round_id, seen


def extract_sheet_rows(xlsx_path: Path) -> dict[str, list[list[object]]]:
    xl = pd.ExcelFile(xlsx_path)
    out: dict[str, list[list[object]]] = {}
    for sheet in xl.sheet_names:
        df = pd.read_excel(xlsx_path, sheet_name=sheet, header=None)
        rows: list[list[object]] = []
        for row in df.itertuples(index=False):
            values = ["" if pd.isna(v) else v for v in row]
            rows.append(values)
        out[sheet] = rows
    return out


def build_combined_round_rows(round_id: str, sheet_rows: dict[str, list[list[object]]]) -> list[list[object]]:
    rows: list[list[object]] = []
    rows.append([f"{round_id.upper()} Market Reports", "", "", "", "", "", "", ""])
    rows.append(["Source", "exschool image OCR verified", "", "", "", "", "", ""])
    rows.append(["", "", "", "", "", "", "", ""])
    for city, city_rows in sheet_rows.items():
        rows.append([f"Market Report - {city}", "", "", "", "", "", "", ""])
        normalized_rows = list(city_rows)
        while normalized_rows and all(str(v).strip() == "" for v in normalized_rows[0]):
            normalized_rows.pop(0)
        if normalized_rows and str(normalized_rows[0][0]).strip().startswith("Market Report - "):
            normalized_rows.pop(0)
        while normalized_rows and all(str(v).strip() == "" for v in normalized_rows[0]):
            normalized_rows.pop(0)
        rows.extend(normalized_rows)
        rows.append(["", "", "", "", "", "", "", ""])
    return rows


def write_html(round_id: str, rows: list[list[object]], city_titles: list[str], output_path: Path) -> None:
    max_cols = max(len(r) for r in rows)
    lines = [
        "<!doctype html>",
        '<html lang="zh-CN">',
        "<head>",
        '  <meta charset="utf-8" />',
        f"  <title>{round_id.upper()} Market Reports</title>",
        "  <style>",
        "    body { font-family: Arial, sans-serif; margin: 24px; }",
        "    table { border-collapse: collapse; width: 100%; }",
        "    td, th { border: 1px solid #999; padding: 6px 8px; font-size: 12px; }",
        "    .title td { background: #1f4e78; color: #fff; font-weight: 700; }",
        "    .section td { background: #d9eaf7; font-weight: 700; }",
        "    .blank td { border: none; height: 8px; }",
        "  </style>",
        "</head>",
        "<body>",
        f"  <h1>{round_id.upper()} OCR Verified Market Reports</h1>",
        f"  <p>OCR detected city sections: {', '.join(city_titles)}</p>",
        "  <table>",
    ]
    for idx, row in enumerate(rows):
        cls = ""
        if idx == 0:
            cls = ' class="title"'
        elif row and isinstance(row[0], str) and str(row[0]).startswith("Market Report - "):
            cls = ' class="section"'
        elif all((str(v).strip() == "" for v in row)):
            cls = ' class="blank"'
        lines.append(f"    <tr{cls}>")
        padded = row + [""] * (max_cols - len(row))
        for value in padded:
            text = "" if value is None else str(value)
            text = (
                text.replace("&", "&amp;")
                .replace("<", "&lt;")
                .replace(">", "&gt;")
            )
            lines.append(f"      <td>{text}</td>")
        lines.append("    </tr>")
    lines += ["  </table>", "</body>", "</html>"]
    output_path.write_text("\n".join(lines), encoding="utf-8")


def write_combined_xlsx_from_html(round_id: str, html_path: Path, output_path: Path) -> None:
    tables = pd.read_html(str(html_path))
    df = tables[0]
    wb = Workbook()
    ws = wb.active
    ws.title = round_id
    title_fill = PatternFill("solid", fgColor="1F4E78")
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    bold = Font(bold=True)
    white_bold = Font(bold=True, color="FFFFFF")
    for row in df.itertuples(index=False):
        ws.append(list(row))
    for cell in ws[1]:
        cell.fill = title_fill
        cell.font = white_bold
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if row[0].value and isinstance(row[0].value, str) and row[0].value.startswith("Market Report - "):
            for cell in row:
                cell.fill = section_fill
                cell.font = bold
    wb.save(output_path)


def write_structured_workbook(round_id: str, sheet_rows: dict[str, list[list[object]]], output_path: Path) -> None:
    wb = Workbook()
    first = True
    for city, rows in sheet_rows.items():
        ws = wb.active if first else wb.create_sheet()
        first = False
        ws.title = city
        for row in rows:
            ws.append(list(row))
    wb.save(output_path)


def validate(round_id: str, image_round: str, ocr_cities: list[str], source_sheets: list[str], structured_path: Path) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    rows.append(
        {
            "round_id": round_id,
            "check": "ocr_round_matches_expected",
            "expected": round_id,
            "actual": image_round,
            "ok": image_round == round_id,
        }
    )
    rows.append(
        {
            "round_id": round_id,
            "check": "ocr_city_titles_match_source_sheets",
            "expected": ",".join(source_sheets),
            "actual": ",".join(ocr_cities),
            "ok": ocr_cities == source_sheets,
        }
    )
    generated_sheets = pd.ExcelFile(structured_path).sheet_names
    rows.append(
        {
            "round_id": round_id,
            "check": "generated_structured_sheets_match_source",
            "expected": ",".join(source_sheets),
            "actual": ",".join(generated_sheets),
            "ok": generated_sheets == source_sheets,
        }
    )
    return rows


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    html_dir = OUTPUT_DIR / "html"
    combined_xlsx_dir = OUTPUT_DIR / "combined_xlsx"
    structured_dir = OUTPUT_DIR / "structured_xlsx"
    html_dir.mkdir(exist_ok=True)
    combined_xlsx_dir.mkdir(exist_ok=True)
    structured_dir.mkdir(exist_ok=True)

    validation_rows: list[dict[str, object]] = []

    for image_name, expected_round in IMAGE_ROUND_MAP.items():
        image_path = EXSCHOOL_DIR / image_name
        image_round, ocr_cities = ocr_image(image_path)

        source_xlsx = ROUND_SOURCE_XLSX[expected_round]
        sheet_rows = extract_sheet_rows(source_xlsx)
        source_sheets = list(sheet_rows.keys())
        combined_rows = build_combined_round_rows(expected_round, sheet_rows)

        html_path = html_dir / f"{expected_round}_market_reports.html"
        combined_xlsx_path = combined_xlsx_dir / f"{expected_round}_market_reports_combined.xlsx"
        structured_xlsx_path = structured_dir / source_xlsx.name

        write_html(expected_round, combined_rows, ocr_cities, html_path)
        write_combined_xlsx_from_html(expected_round, html_path, combined_xlsx_path)
        write_structured_workbook(expected_round, sheet_rows, structured_xlsx_path)

        validation_rows.extend(validate(expected_round, image_round, ocr_cities, source_sheets, structured_xlsx_path))

    validation_df = pd.DataFrame(validation_rows)
    validation_path = OUTPUT_DIR / "validation_report.xlsx"
    validation_df.to_excel(validation_path, index=False)

    print(f"Wrote HTML and Excel outputs to {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
