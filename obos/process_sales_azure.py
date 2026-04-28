#!/usr/bin/env python3
import os
import re
from pathlib import Path

from azure.ai.formrecognizer import DocumentAnalysisClient
from azure.core.credentials import AzureKeyCredential
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path("/mnt/c/Users/david/documents/ASDAN/表格/WYEF")
OUTPUT_DIR = Path("/mnt/c/Users/david/documents/ASDAN/表格/WYEF_results")
ROUND_ORDER = ["r-1", "r1", "r2", "r3", "r4", "r5", "r6", "r7"]
CANONICAL_MARKETS = [
    "Shanghai",
    "Seoul",
    "Montreal",
    "Paris",
    "Tokyo",
    "Stockholm",
    "London",
    "Berlin",
    "Sydney",
    "San Francisco",
]

AGENTS_HEADERS = [
    "Market",
    "Previous",
    "Change",
    "After",
    "Change Cost",
    "Marketing Investment",
]
MARKET_HEADERS = [
    "Market",
    "Competitive Power",
    "Sales Volume",
    "Market Share",
    "Price",
    "Sales",
]


def get_client():
    endpoint = os.environ.get("DOCUMENTINTELLIGENCE_ENDPOINT")
    key = os.environ.get("DOCUMENTINTELLIGENCE_API_KEY")
    if not endpoint or not key:
        raise RuntimeError(
            "DOCUMENTINTELLIGENCE_ENDPOINT or DOCUMENTINTELLIGENCE_API_KEY is missing"
        )
    return DocumentAnalysisClient(
        endpoint=endpoint,
        credential=AzureKeyCredential(key),
    )


def find_sales_image(round_dir: Path):
    candidates = [round_dir / "sales.png", round_dir / f"{round_dir.name}sales.png"]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def analyze_image(client, image_path: Path):
    with image_path.open("rb") as fh:
        poller = client.begin_analyze_document("prebuilt-layout", document=fh)
    return poller.result()


def clean_text(value):
    if value is None:
        return ""
    text = str(value).replace("\n", " ").strip()
    return re.sub(r"\s+", " ", text)


def parse_int(value):
    text = clean_text(value)
    if not text:
        return None
    digits = re.sub(r"[^\d-]", "", text)
    if not digits or digits == "-":
        return None
    return int(digits)


def parse_signed_int(value):
    text = clean_text(value)
    if not text:
        return None
    sign = -1 if "-" in text and "+" not in text else 1
    digits = re.sub(r"[^\d]", "", text)
    if not digits:
        return None
    return sign * int(digits)


def parse_percent(value):
    text = clean_text(value).replace(",", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    return float(match.group(0)) / 100.0


def table_to_rows(table):
    rows = [["" for _ in range(table.column_count)] for _ in range(table.row_count)]
    for cell in table.cells:
        rows[cell.row_index][cell.column_index] = clean_text(cell.content)
    return rows


def find_agents_table(result):
    for table in result.tables:
        rows = table_to_rows(table)
        if not rows:
            continue
        header_text = " ".join(rows[0]).lower()
        if "agents" in header_text and "previous" in header_text:
            return rows
    raise RuntimeError("Agents table not found")


def collapse_agents_row(row):
    row = [clean_text(cell) for cell in row]
    if len(row) > 6:
        row = row[:5] + [" ".join(part for part in row[5:] if part)]
    if len(row) < 6:
        row = row + [""] * (6 - len(row))
    return row[:6]


def normalize_agents_rows(table_rows):
    data_rows = [collapse_agents_row(row) for row in table_rows[1:] if any(clean_text(v) for v in row)]
    if len(data_rows) != len(CANONICAL_MARKETS):
        raise RuntimeError(
            f"Agents rows mismatch: expected {len(CANONICAL_MARKETS)}, got {len(data_rows)}"
        )

    normalized = []
    for idx, row in enumerate(data_rows):
        previous = parse_int(row[1])
        change = parse_signed_int(row[2])
        after = parse_int(row[3])
        change_cost = parse_int(row[4])
        marketing_investment = parse_int(row[5])

        if change is None and previous is not None and after is not None:
            change = after - previous
        if change is None and change_cost == 0:
            change = 0
        if after is None and previous is not None and change is not None:
            after = previous + change
        if after is None and previous is not None and change_cost == 0:
            after = previous
        if change_cost is None and change is not None and change >= 0:
            change_cost = change * 300000
        if marketing_investment is None and row[5] in {"", "¥0", "0"}:
            marketing_investment = 0

        normalized.append(
            {
                "Market": CANONICAL_MARKETS[idx],
                "Previous": previous if previous is not None else 0,
                "Change": change if change is not None else 0,
                "After": after if after is not None else 0,
                "Change Cost": change_cost if change_cost is not None else 0,
                "Marketing Investment": marketing_investment if marketing_investment is not None else 0,
            }
        )
    return normalized


def extract_lines(result):
    lines = []
    for page in result.pages:
        for line in page.lines:
            text = clean_text(line.content)
            if text:
                lines.append(text)
    return lines


def find_sequence(lines, sequence):
    for idx in range(len(lines) - len(sequence) + 1):
        if lines[idx : idx + len(sequence)] == sequence:
            return idx
    return -1


def normalize_market_name(name, fallback_idx):
    text = clean_text(name).lower()
    alias_map = {
        "san": "San Francisco",
        "san francisco": "San Francisco",
    }
    if text in alias_map:
        return alias_map[text]
    if 0 <= fallback_idx < len(CANONICAL_MARKETS):
        return CANONICAL_MARKETS[fallback_idx]
    return clean_text(name)


def parse_market_rows_from_lines(lines):
    required_prefix = MARKET_HEADERS[:4]
    start = find_sequence(lines, required_prefix)
    if start == -1:
        raise RuntimeError("Market section not found in OCR lines")

    cursor = start + len(required_prefix)
    while cursor < len(lines) and lines[cursor] in {"Price", "Sales"}:
        cursor += 1

    tokens = lines[cursor:]
    expected_rows = len(CANONICAL_MARKETS)
    rows = []
    token_idx = 0

    while token_idx + 5 < len(tokens) and len(rows) < expected_rows:
        chunk = tokens[token_idx : token_idx + 6]
        rows.append(chunk)
        token_idx += 6

    if len(rows) != expected_rows:
        raise RuntimeError(
            f"Market rows mismatch: expected {expected_rows}, got {len(rows)}"
        )
    return rows


def normalize_market_rows(line_rows):
    normalized = []
    for idx, row in enumerate(line_rows):
        market_name = normalize_market_name(row[0], idx)
        competitive_power = parse_percent(row[1])
        sales_volume = parse_int(row[2])
        market_share = parse_percent(row[3])
        price = parse_int(row[4])
        sales = parse_int(row[5])

        if sales_volume is None and sales is not None and price not in (None, 0):
            sales_volume = sales // price
        if sales is not None and price not in (None, 0) and sales % price == 0:
            sales_volume = sales // price
        if sales is None and sales_volume is not None and price is not None:
            sales = sales_volume * price

        if sales_volume is not None and price is not None:
            computed_sales = sales_volume * price
            if sales is None or sales != computed_sales:
                sales = computed_sales

        normalized.append(
            {
                "Market": market_name,
                "Competitive Power": competitive_power if competitive_power is not None else 0,
                "Sales Volume": sales_volume if sales_volume is not None else 0,
                "Market Share": market_share if market_share is not None else 0,
                "Price": price if price is not None else 0,
                "Sales": sales if sales is not None else 0,
            }
        )
    return normalized


def style_sheet(ws):
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_alignment = Alignment(horizontal="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment

    ws.freeze_panes = "A2"

    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_len = max(len(clean_text(cell.value)) for cell in column_cells if cell.value is not None)
        ws.column_dimensions[column_letter].width = min(max_len + 2, 24)


def write_agents_sheet(wb, rows):
    ws = wb.create_sheet("Agents")
    ws.append(AGENTS_HEADERS)
    for row in rows:
        ws.append([row[header] for header in AGENTS_HEADERS])

    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row_idx, 2).number_format = "#,##0"
        ws.cell(row_idx, 3).number_format = "+0;-0;0"
        ws.cell(row_idx, 4).number_format = "#,##0"
        ws.cell(row_idx, 5).number_format = "#,##0"
        ws.cell(row_idx, 6).number_format = "#,##0"

    style_sheet(ws)


def write_market_sheet(wb, rows):
    ws = wb.create_sheet("Market Sales")
    ws.append(MARKET_HEADERS)
    for row in rows:
        ws.append([row[header] for header in MARKET_HEADERS])

    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row_idx, 2).number_format = "0.00%"
        ws.cell(row_idx, 3).number_format = "#,##0"
        ws.cell(row_idx, 4).number_format = "0.00%"
        ws.cell(row_idx, 5).number_format = "#,##0"
        ws.cell(row_idx, 6).number_format = "#,##0"

    style_sheet(ws)


def save_workbook(round_name, agents_rows, market_rows):
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_path = OUTPUT_DIR / f"{round_name}_compeptiveindex_summary.xlsx"

    wb = Workbook()
    wb.remove(wb.active)

    write_agents_sheet(wb, agents_rows)
    write_market_sheet(wb, market_rows)
    wb.save(output_path)
    return output_path


def process_round(client, round_name):
    round_dir = BASE_DIR / round_name
    image_path = find_sales_image(round_dir)
    if image_path is None:
        print(f"[skip] {round_name}: sales image not found")
        return None

    print(f"[run ] {round_name}: {image_path.name}")
    result = analyze_image(client, image_path)
    agents_table = find_agents_table(result)
    lines = extract_lines(result)

    agents_rows = normalize_agents_rows(agents_table)
    market_line_rows = parse_market_rows_from_lines(lines)
    market_rows = normalize_market_rows(market_line_rows)
    output_path = save_workbook(round_name, agents_rows, market_rows)

    print(f"[done] {round_name}: {output_path.name}")
    return output_path


def main():
    client = get_client()
    generated = []

    for round_name in ROUND_ORDER:
        output = process_round(client, round_name)
        if output is not None:
            generated.append(output)

    print(f"\nGenerated {len(generated)} workbook(s).")
    for output in generated:
        print(output)


if __name__ == "__main__":
    main()
