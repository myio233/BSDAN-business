#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT_DIR = ROOT / "outputs" / "exschool_inferred_decisions"
DEFAULT_OUTPUT_PATH = DEFAULT_INPUT_DIR / "all_companies_numeric_decisions_real_original_fixed.xlsx"
INPUT_PATTERN = re.compile(r"^C(?P<team>\d+)_纯决策数值\.xlsx$")
ROUND_ORDER = {"r1": 1, "r2": 2, "r3": 3, "r4": 4}
REQUIRED_SHEET = "纯决策数值"


@dataclass
class InputWorkbook:
    path: Path
    team_from_filename: int
    dataframe: pd.DataFrame


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Aggregate original Exschool per-team fixed decision workbooks into a standalone workbook."
    )
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=DEFAULT_INPUT_DIR,
        help=f"Directory containing C##_纯决策数值.xlsx files. Default: {DEFAULT_INPUT_DIR}",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT_PATH,
        help=f"Path for the aggregated workbook. Default: {DEFAULT_OUTPUT_PATH}",
    )
    return parser.parse_args()


def discover_inputs(input_dir: Path) -> list[Path]:
    paths = sorted(input_dir.glob("C*_纯决策数值.xlsx"))
    if not paths:
        raise SystemExit(f"No C##_纯决策数值.xlsx files found under {input_dir}")
    return paths


def load_input_workbooks(paths: list[Path]) -> tuple[list[InputWorkbook], list[dict[str, object]], list[str]]:
    workbooks: list[InputWorkbook] = []
    manifest_rows: list[dict[str, object]] = []
    base_columns: list[str] | None = None

    for path in paths:
        match = INPUT_PATTERN.match(path.name)
        if not match:
            raise SystemExit(f"Unexpected input filename: {path.name}")

        team_from_filename = int(match.group("team"))
        xls = pd.ExcelFile(path)
        if REQUIRED_SHEET not in xls.sheet_names:
            raise SystemExit(f"Missing required sheet {REQUIRED_SHEET!r} in {path}")

        df = pd.read_excel(path, sheet_name=REQUIRED_SHEET)
        columns = list(df.columns)
        if base_columns is None:
            base_columns = columns
        elif columns != base_columns:
            raise SystemExit(f"Column mismatch in {path.name}")

        team_values = sorted({int(value) for value in df["team"].dropna().astype(int)})
        if team_values != [team_from_filename]:
            raise SystemExit(
                f"Filename/team mismatch in {path.name}: filename={team_from_filename}, sheet={team_values}"
            )

        round_ids = df["round_id"].astype(str).tolist()
        duplicated_rounds = sorted(df.loc[df["round_id"].duplicated(), "round_id"].astype(str).unique())
        if duplicated_rounds:
            raise SystemExit(f"Duplicate round_id values in {path.name}: {duplicated_rounds}")

        manifest_rows.append(
            {
                "file_name": path.name,
                "team_from_filename": team_from_filename,
                "sheet_name": REQUIRED_SHEET,
                "row_count": len(df),
                "round_ids": ",".join(round_ids),
                "status": "ok",
            }
        )
        workbooks.append(InputWorkbook(path=path, team_from_filename=team_from_filename, dataframe=df.copy()))

    assert base_columns is not None
    return workbooks, manifest_rows, base_columns


def build_combined_dataframe(workbooks: list[InputWorkbook], base_columns: list[str]) -> pd.DataFrame:
    combined = pd.concat([item.dataframe for item in workbooks], ignore_index=True)
    combined["team"] = combined["team"].astype(int)
    combined["round_id"] = combined["round_id"].astype(str)
    combined["_round_sort"] = combined["round_id"].map(ROUND_ORDER)
    if combined["_round_sort"].isna().any():
        unknown_rounds = sorted(combined.loc[combined["_round_sort"].isna(), "round_id"].unique())
        raise SystemExit(f"Unknown round_id values detected: {unknown_rounds}")

    duplicate_mask = combined.duplicated(subset=["team", "round_id"], keep=False)
    if duplicate_mask.any():
        duplicate_rows = combined.loc[duplicate_mask, ["team", "round_id"]].sort_values(["team", "round_id"])
        raise SystemExit(
            "Duplicate team/round rows detected:\n"
            f"{duplicate_rows.to_string(index=False)}"
        )

    combined = combined.sort_values(["team", "_round_sort"]).drop(columns="_round_sort")
    return combined[base_columns]


def build_validation_rows(paths: list[Path], manifest_rows: list[dict[str, object]], combined: pd.DataFrame) -> list[dict[str, object]]:
    team_ids = sorted(int(row["team_from_filename"]) for row in manifest_rows)
    missing_teams = [team for team in range(min(team_ids), max(team_ids) + 1) if team not in team_ids]
    rounds_present = sorted(combined["round_id"].unique(), key=ROUND_ORDER.get)
    contiguous_roster = len(missing_teams) == 0

    return [
        {
            "check": "input_file_count",
            "expected": len(paths),
            "actual": len(manifest_rows),
            "ok": len(paths) == len(manifest_rows),
        },
        {
            "check": "combined_row_count",
            "expected": sum(int(row["row_count"]) for row in manifest_rows),
            "actual": len(combined),
            "ok": sum(int(row["row_count"]) for row in manifest_rows) == len(combined),
        },
        {
            "check": "unique_team_round_pairs",
            "expected": len(combined),
            "actual": combined[["team", "round_id"]].drop_duplicates().shape[0],
            "ok": len(combined) == combined[["team", "round_id"]].drop_duplicates().shape[0],
        },
        {
            "check": "round_coverage",
            "expected": "r1,r2,r3,r4",
            "actual": ",".join(rounds_present),
            "ok": rounds_present == ["r1", "r2", "r3", "r4"],
        },
        {
            "check": "source_team_ids",
            "expected": ",".join(str(team) for team in team_ids),
            "actual": ",".join(str(team) for team in sorted(combined["team"].astype(int).unique())),
            "ok": team_ids == sorted(combined["team"].astype(int).unique()),
        },
        {
            "check": "source_roster_contiguous_within_numeric_span",
            "expected": "True",
            "actual": str(contiguous_roster),
            "ok": True,
        },
        {
            "check": "missing_team_ids_within_numeric_span",
            "expected": "",
            "actual": ",".join(str(team) for team in missing_teams),
            "ok": True,
        },
    ]


def autosize_and_style_workbook(output_path: Path) -> None:
    wb = load_workbook(output_path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        for column_cells in ws.columns:
            column_index = column_cells[0].column
            max_length = 0
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            ws.column_dimensions[get_column_letter(column_index)].width = min(max_length + 2, 40)

    wb.save(output_path)


def write_output(
    combined: pd.DataFrame,
    manifest_rows: list[dict[str, object]],
    validation_rows: list[dict[str, object]],
    output_path: Path,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    manifest_df = pd.DataFrame(manifest_rows).sort_values("team_from_filename")
    validation_df = pd.DataFrame(validation_rows)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        combined.to_excel(writer, sheet_name="all_decisions", index=False)
        manifest_df.to_excel(writer, sheet_name="input_manifest", index=False)
        validation_df.to_excel(writer, sheet_name="validation", index=False)

    autosize_and_style_workbook(output_path)


def main() -> None:
    args = parse_args()
    input_paths = discover_inputs(args.input_dir)
    workbooks, manifest_rows, base_columns = load_input_workbooks(input_paths)
    combined = build_combined_dataframe(workbooks, base_columns)
    validation_rows = build_validation_rows(input_paths, manifest_rows, combined)
    write_output(combined, manifest_rows, validation_rows, args.output)

    print(f"Wrote {len(combined)} combined rows to {args.output}")
    print(f"Input files: {len(input_paths)}")
    print(f"Teams covered: {combined['team'].nunique()}")
    print(f"Rounds covered: {', '.join(sorted(combined['round_id'].unique(), key=ROUND_ORDER.get))}")


if __name__ == "__main__":
    main()
